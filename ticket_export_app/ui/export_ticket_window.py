# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (
    QApplication, QDialog, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QToolBar, QStatusBar, QMessageBox, QTableWidget,
    QTableWidgetItem, QSpinBox, QComboBox, QLineEdit, QColorDialog,
    QTabWidget, QFrame, QAbstractItemView, QHeaderView, QGraphicsScene, QGraphicsView, QProgressBar,
    QPlainTextEdit
)
from PySide6.QtCore import Qt, QThreadPool, QTimer
from PySide6.QtGui import QAction, QColor, QPen, QBrush, QFont
import os
import math
from html import escape
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill, Border, Side

# 版本号：优先从当前工程的 __init__ 里取，取不到就用 "dev"
try:
    from __init__ import __version__
except Exception:
    __version__ = "dev"

# Worker & tickets：直接从当前工程内部模块导入
from infra.threads import Worker
from core import tickets
from core.analysis import apply_time_window_analysis as core_apply_time_window_analysis
from core.input_parser import parse_multi_project_inputs as core_parse_multi_project_inputs


class ExportTicketWindow(QMainWindow):
    """
    导出组合票（独立于数据校对）
    v2 岗位矩阵字段：
      序号 / 工程名称 / 设备数量 / 所属线别 / 岗位设备 / A工时 / B工时 / C工时
    说明：
      - 设备数量：1 表示单资源；2 表示双线双资源。
      - 所属线别：1号线 / 2号线 / 双线 / 双线共用。
      - A/B/C 工时 > 0：该车型经过该岗位；工时 = 0：该车型跳过该岗位。
      - 参与投车的车型，工时不能为空；未参与投车的车型，工时可以为空。
    """
    COL_C_TIME = 7
    MAX_SINGLE_STEPS = 23  # 单工程组合票：新版模板固定支持 23 行

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"导出组合票  v{__version__}")
        self.resize(1180, 760)
        self.setMinimumSize(1080, 700)

        self.thread_pool = QThreadPool.globalInstance()
        self.dst_path = None

        self._build_ui()
        self._connect_signals()
    def _on_tab_changed(self, index: int):
        """
        Tab 切换时，控制多工程页内“添加步骤 / 删除步骤 / 填入示例”按钮：
        - 仅在『多工程组合票』页签（第 0 个 Tab）启用；
        - 在『单工程组合票』页签禁用，避免误点影响单工程表。
        """
        is_multi = (index == 0)
        if hasattr(self, "btn_add_row"):
            self.btn_add_row.setEnabled(is_multi)
        if hasattr(self, "btn_del_row"):
            self.btn_del_row.setEnabled(is_multi)
        if hasattr(self, "btn_fill_sample"):
            self.btn_fill_sample.setEnabled(is_multi)
    # ---------------- UI ---------------- #
    def _build_ui(self):
        tb = QToolBar("Ticket")
        self.addToolBar(tb)

        self.act_help = QAction("帮助", self)
        tb.addAction(self.act_help)
        tb.addSeparator()

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        # ====== Tab 控件 ======
        self.tabs = QTabWidget(self)
        root.addWidget(self.tabs)

        # ---------- Tab1：多车组合票 ----------
        self.page_multi = QWidget(self)
        page_multi_layout = QVBoxLayout(self.page_multi)
        page_multi_layout.setContentsMargins(8, 8, 8, 8)
        page_multi_layout.setSpacing(10)
        self.tabs.addTab(self.page_multi, "多工程组合票")

        # 多工程内部二级页：第一页负责录入，第二页负责分析/导出/后续动画预留
        self.multi_tabs = QTabWidget(self.page_multi)
        page_multi_layout.addWidget(self.multi_tabs, 1)

        self.page_multi_input = QWidget(self.page_multi)
        self.page_multi_result = QWidget(self.page_multi)

        input_layout = QVBoxLayout(self.page_multi_input)
        input_layout.setContentsMargins(0, 0, 0, 0)
        input_layout.setSpacing(10)

        result_layout = QVBoxLayout(self.page_multi_result)
        result_layout.setContentsMargins(0, 0, 0, 0)
        result_layout.setSpacing(10)


        self.multi_tabs.addTab(self.page_multi_input, "参数与岗位")
        self.multi_tabs.addTab(self.page_multi_result, "分析与导出")


        def _make_block(title: str):
            frame = QFrame(self.page_multi)
            frame.setFrameShape(QFrame.StyledPanel)
            frame.setObjectName("ticketBlock")
            frame.setStyleSheet(
                "QFrame#ticketBlock {"
                "background: #ffffff;"
                "border: 1px solid #d9dee7;"
                "border-radius: 10px;"
                "}"
            )
            lay = QVBoxLayout(frame)
            lay.setContentsMargins(12, 12, 12, 12)
            lay.setSpacing(10)
            title_label = QLabel(title, frame)
            title_label.setStyleSheet("font-size: 14px; font-weight: 700; color: #223042;")
            lay.addWidget(title_label)
            return frame, lay

        top_split = QVBoxLayout()
        top_split.setSpacing(10)
        input_layout.addLayout(top_split)

        # 左侧：运行参数区块
        params_frame, params_layout = _make_block("运行参数")
        top_split.addWidget(params_frame, 0)

        row_top_1 = QHBoxLayout()
        row_top_1.setSpacing(8)
        params_layout.addLayout(row_top_1)
        row_top_1.addWidget(QLabel("工程名称："))
        self.ed_project = QLineEdit()
        self.ed_project.setPlaceholderText("例如：L2++")
        self.ed_project.setFixedWidth(220)
        row_top_1.addWidget(self.ed_project)

        row_top_1.addSpacing(16)
        row_top_1.addWidget(QLabel("投车模式："))
        self.cmb_launch_mode = QComboBox()
        self.cmb_launch_mode.addItems(["按数量投车", "按比例投车"])
        self.cmb_launch_mode.setCurrentIndex(0)
        row_top_1.addWidget(self.cmb_launch_mode)
        row_top_1.addStretch()

        row_top_2 = QHBoxLayout()
        row_top_2.setSpacing(8)
        params_layout.addLayout(row_top_2)
        self.lbl_a_cars = QLabel("A数量：")
        row_top_2.addWidget(self.lbl_a_cars)
        self.spn_a_cars = QSpinBox()
        self.spn_a_cars.setRange(0, 9999)
        self.spn_a_cars.setValue(4)
        row_top_2.addWidget(self.spn_a_cars)

        row_top_2.addSpacing(10)
        self.lbl_b_cars = QLabel("B数量：")
        row_top_2.addWidget(self.lbl_b_cars)
        self.spn_b_cars = QSpinBox()
        self.spn_b_cars.setRange(0, 9999)
        self.spn_b_cars.setValue(0)
        row_top_2.addWidget(self.spn_b_cars)

        row_top_2.addSpacing(10)
        self.lbl_c_cars = QLabel("C数量：")
        row_top_2.addWidget(self.lbl_c_cars)
        self.spn_c_cars = QSpinBox()
        self.spn_c_cars.setRange(0, 9999)
        self.spn_c_cars.setValue(0)
        row_top_2.addWidget(self.spn_c_cars)
        row_top_2.addStretch()

        row_ratio = QHBoxLayout()
        row_ratio.setSpacing(8)
        params_layout.addLayout(row_ratio)
        self.lbl_total_cars = QLabel("分析时间：")
        row_ratio.addWidget(self.lbl_total_cars)
        self.spn_total_cars = QSpinBox()
        self.spn_total_cars.setRange(1, 9999)
        self.spn_total_cars.setValue(60)
        self.spn_total_cars.setSuffix(" 分钟")
        self.spn_total_cars.setToolTip(
            "按比例投车模式下使用；空线起步，第1台从0s进入首岗位，按分析时间和目标节拍计算理论投车台数。"
        )
        row_ratio.addWidget(self.spn_total_cars)
        self.lbl_total_cars.hide()
        self.spn_total_cars.hide()

        # 旧比例文本框保留但隐藏，后续可删除；当前按比例模式改用 A/B/C 数值框作为比例。
        self.ed_ratio = QLineEdit()
        self.ed_ratio.hide()

        row_ratio.addStretch()

        row_top_3 = QHBoxLayout()
        row_top_3.setSpacing(8)
        params_layout.addLayout(row_top_3)
        row_top_3.addWidget(QLabel("时间格刻度："))
        self.cmb_grid = QComboBox()
        self.cmb_grid.addItems(["1.0", "0.5", "2.0"])
        self.cmb_grid.setCurrentIndex(0)
        row_top_3.addWidget(self.cmb_grid)

        # 等待分配先隐藏，底层默认按“开始前等待”
        self.cmb_wait = QComboBox()
        self.cmb_wait.addItems(["开始前等待", "末尾等待"])
        self.cmb_wait.setCurrentIndex(0)
        self.cmb_wait.hide()

        row_top_3.addSpacing(12)
        row_top_3.addWidget(QLabel("目标节拍："))
        self.spn_target_takt = QSpinBox()
        self.spn_target_takt.setRange(0, 9999)
        self.spn_target_takt.setValue(118)
        self.spn_target_takt.setToolTip("0表示不进行节拍判定；大于0时按各岗位A/B/C实际工时判断OK/NG")
        row_top_3.addWidget(self.spn_target_takt)
        
        row_top_3.addSpacing(12)
        self.lbl_sequence_mode = QLabel("排列方式：")
        row_top_3.addWidget(self.lbl_sequence_mode)
        self.cmb_seq = QComboBox()
        self.cmb_seq.addItems(["顺排(A→B→C)", "交替混流"])
        self.cmb_seq.setCurrentIndex(0)
        row_top_3.addWidget(self.cmb_seq)

        row_top_3.addSpacing(12)
        self.lbl_max_run = QLabel("最大连续台数：")
        row_top_3.addWidget(self.lbl_max_run)
        self.spn_max_run = QSpinBox()
        self.spn_max_run.setRange(1, 9999)
        self.spn_max_run.setValue(10)
        self.spn_max_run.setToolTip("默认10台；填1表示尽量强制交替")
        row_top_3.addWidget(self.spn_max_run)
        row_top_3.addStretch()

        self.params_tip = QLabel(
            "顺排/交替混流：A/B/C 填数量；按比例投车：A/B/C 填比例，并填写分析时间。下方岗位矩阵用于逐行录入步骤。"
        )
        self.params_tip.setWordWrap(True)
        self.params_tip.setStyleSheet("color: #5f6b7a; font-size: 12px;")
        params_layout.addWidget(self.params_tip)

        # 中部：岗位矩阵区块
        table_frame, table_layout = _make_block("岗位矩阵")

        table_action_row = QHBoxLayout()
        table_action_row.setSpacing(8)
        table_layout.addLayout(table_action_row)
        table_action_row.addStretch()
        self.btn_add_row = QPushButton("添加步骤", self.page_multi_input)
        self.btn_del_row = QPushButton("删除步骤", self.page_multi_input)
        self.btn_fill_sample = QPushButton("填入示例", self.page_multi_input)
        table_action_row.addWidget(self.btn_add_row)
        table_action_row.addWidget(self.btn_del_row)
        table_action_row.addWidget(self.btn_fill_sample)

        input_layout.addWidget(table_frame, 1)

        input_next_row = QHBoxLayout()
        input_next_row.setSpacing(8)
        input_next_row.addStretch()
        self.btn_export = QPushButton("生成并导出组合票", self.page_multi_input)
        input_next_row.addWidget(self.btn_export)
        self.btn_go_result_page = QPushButton("下一步：分析与导出", self.page_multi_input)
        input_next_row.addWidget(self.btn_go_result_page)
        input_layout.addLayout(input_next_row)

        self.tbl = QTableWidget(0, 8, self)
        self.tbl.setHorizontalHeaderLabels([
            "序号", "工程名称", "设备数量", "所属线别",
            "岗位设备", "A工时", "B工时", "C工时"
        ])
        self.tbl.horizontalHeader().setStretchLastSection(False)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
        )
        self.tbl.setTabKeyNavigation(False)
        self.tbl.setAlternatingRowColors(True)
        table_layout.addWidget(self.tbl, 1)

        # 分析页顶部：只保留一个小按钮行，避免占用纵向空间。
        result_nav_row = QHBoxLayout()
        result_nav_row.setSpacing(8)
        result_nav_row.addStretch()
        self.btn_analyze = QPushButton("分析当前排程", self.page_multi_result)
        result_nav_row.addWidget(self.btn_analyze)
        result_layout.addLayout(result_nav_row)

        # 以下控件仅作为旧逻辑兼容容器保留，不再占用分析页版面空间。
        self.lbl_analysis = QLabel(
            "结果分析：点击『分析当前排程』后显示总车数、总完成时间、总等待时间、平均等待时间与节拍判定。",
            self.page_multi_result,
        )
        self.lbl_analysis.setWordWrap(True)
        self.lbl_analysis.hide()

        self.tbl_station_analysis = QTableWidget(0, 8, self.page_multi_result)
        self.tbl_station_analysis.setHorizontalHeaderLabels([
            "岗位", "经过台数", "累计工时", "累计等待", "平均工时", "平均等待", "节拍判定", "超节拍车型"
        ])
        self.tbl_station_analysis.horizontalHeader().setStretchLastSection(False)
        self.tbl_station_analysis.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_station_analysis.verticalHeader().setVisible(False)
        self.tbl_station_analysis.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_station_analysis.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_station_analysis.setAlternatingRowColors(True)
        self.tbl_station_analysis.hide()

        # 车型数据摘要区：只承载基础排程摘要与模型结果。
        vehicle_summary_frame, vehicle_summary_layout = _make_block("车型数据摘要区")
        result_layout.addWidget(vehicle_summary_frame, 3)

        self.sim_timer = QTimer(self)
        self.sim_timer.setInterval(100)
        self.sim_time = 0.0
        self.current_defs = []
        self.last_schedule_rows = []
        self.last_analysis = None
        self.last_max_finish = 0.0

        self.lbl_vehicle_summary = QLabel("")
        self.lbl_vehicle_summary.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.lbl_vehicle_summary.setWordWrap(True)
        self.lbl_vehicle_summary.setTextFormat(Qt.RichText)
        self.lbl_vehicle_summary.setMinimumHeight(120)
        self.lbl_vehicle_summary.setMaximumHeight(155)
        self.lbl_vehicle_summary.setStyleSheet(
            "background: #f7f9fc;"
            "border: 1px dashed #cbd5e1;"
            "border-radius: 8px;"
            "padding: 10px;"
            "color: #334155;"
            "font-size: 13px;"
            "line-height: 1.5;"
        )

        vehicle_summary_layout.addWidget(self.lbl_vehicle_summary, 1)

        simulation_frame = QFrame(self.page_multi)
        simulation_frame.setFrameShape(QFrame.StyledPanel)
        simulation_frame.setObjectName("ticketBlock")
        simulation_frame.setStyleSheet(
            "QFrame#ticketBlock {"
            "background: #ffffff;"
            "border: 1px solid #d9dee7;"
            "border-radius: 10px;"
            "}"
        )
        simulation_layout = QVBoxLayout(simulation_frame)
        simulation_layout.setContentsMargins(12, 12, 12, 10)
        simulation_layout.setSpacing(6)
        result_layout.addWidget(simulation_frame, 7)

        sim_control_row = QHBoxLayout()
        sim_control_row.setContentsMargins(0, 0, 0, 0)
        sim_control_row.setSpacing(6)
        sim_title = QLabel("仿真回放", simulation_frame)
        sim_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #223042;")
        sim_control_row.addWidget(sim_title)
        sim_control_row.addStretch(1)
        self.lbl_sim_time = QLabel("仿真时间：0.0s / 0.0s")
        sim_control_row.addWidget(self.lbl_sim_time)
        sim_control_row.addSpacing(10)
        self.lbl_sim_total_wait = QLabel("总等待：0s")
        self.lbl_sim_total_wait.hide()
        sim_control_row.addSpacing(10)
        self.sim_progress = QProgressBar(self.page_multi_result)
        self.sim_progress.setRange(0, 1000)
        self.sim_progress.setValue(0)
        self.sim_progress.setTextVisible(True)
        self.sim_progress.setMinimumWidth(180)
        self.sim_progress.setMaximumWidth(260)
        self.sim_progress.setMaximumHeight(18)
        self.sim_progress.setFormat("进度 %p%")
        sim_control_row.addWidget(self.sim_progress)
        self.cmb_sim_speed = QComboBox()
        self.cmb_sim_speed.addItems(["1x", "5x", "10x", "50x"])
        self.cmb_sim_speed.setCurrentText("10x")
        self.btn_sim_play = QPushButton("播放")
        self.btn_sim_pause = QPushButton("暂停")
        self.btn_sim_reset = QPushButton("重置")
        control_h = 26
        self.cmb_sim_speed.setMinimumHeight(control_h)
        self.cmb_sim_speed.setMaximumHeight(control_h)
        self.btn_sim_play.setMinimumHeight(control_h)
        self.btn_sim_play.setMaximumHeight(control_h)
        self.btn_sim_pause.setMinimumHeight(control_h)
        self.btn_sim_pause.setMaximumHeight(control_h)
        self.btn_sim_reset.setMinimumHeight(control_h)
        self.btn_sim_reset.setMaximumHeight(control_h)
        self.btn_sim_play.setMinimumWidth(52)
        self.btn_sim_pause.setMinimumWidth(52)
        self.btn_sim_reset.setMinimumWidth(52)
        sim_control_row.addWidget(QLabel("速度："))
        sim_control_row.addWidget(self.cmb_sim_speed)
        sim_control_row.addWidget(self.btn_sim_play)
        sim_control_row.addWidget(self.btn_sim_pause)
        sim_control_row.addWidget(self.btn_sim_reset)

        self.sim_scene = QGraphicsScene(self)
        self.sim_graphics_view = QGraphicsView(self.sim_scene, simulation_frame)
        self.sim_graphics_view.setMinimumHeight(232)
        self.sim_graphics_view.setMaximumHeight(252)
        self.sim_graphics_view.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.sim_graphics_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.sim_graphics_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sim_graphics_view.setStyleSheet(
            "background: #243244;"
            "border: 1px solid #334155;"
            "border-radius: 8px;"
        )
        simulation_layout.addLayout(sim_control_row, 0)
        simulation_layout.addWidget(self.sim_graphics_view, 0)

        sim_status_frame = QFrame(simulation_frame)
        sim_status_frame.setObjectName("simStatusFrame")
        sim_status_frame.setStyleSheet(
            "QFrame#simStatusFrame {"
            "background: #f8fafc;"
            "border: 1px solid #cbd5e1;"
            "border-radius: 6px;"
            "}"
        )
        sim_status_layout = QVBoxLayout(sim_status_frame)
        sim_status_layout.setContentsMargins(10, 6, 10, 5)
        sim_status_layout.setSpacing(0)

        self.lbl_sim_view = QLabel("仿真画面：请先点击『分析当前排程』。", sim_status_frame)
        self.lbl_sim_view.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.lbl_sim_view.setWordWrap(True)
        self.lbl_sim_view.setTextFormat(Qt.RichText)
        self.lbl_sim_view.setMinimumHeight(52)
        self.lbl_sim_view.setMaximumHeight(58)
        self.lbl_sim_view.setStyleSheet(
            "background: transparent;"
            "border: 0;"
            "padding: 0;"
            "color: #334155;"
            "font-size: 12px;"
        )
        sim_status_layout.addWidget(self.lbl_sim_view)
        sim_status_frame.setMinimumHeight(66)
        sim_status_frame.setMaximumHeight(72)
        simulation_layout.addStretch(1)
        simulation_layout.addWidget(sim_status_frame, 0)

        vehicle_log_entry = QFrame(self.page_multi_result)
        vehicle_log_entry.setMaximumHeight(44)
        vehicle_log_entry.setStyleSheet(
            "background: #ffffff;"
            "border: 1px solid #e2e8f0;"
            "border-radius: 8px;"
        )
        vehicle_log_layout = QHBoxLayout(vehicle_log_entry)
        vehicle_log_layout.setContentsMargins(12, 8, 12, 8)
        vehicle_log_layout.setSpacing(8)
        vehicle_log_title = QLabel("车辆明细 / 调试日志", vehicle_log_entry)
        vehicle_log_title.setStyleSheet("font-size: 13px; font-weight: 700; color: #334155;")
        vehicle_log_layout.addWidget(vehicle_log_title)
        vehicle_log_layout.addStretch()
        self.btn_vehicle_log = QPushButton("查看车辆日志", vehicle_log_entry)
        vehicle_log_layout.addWidget(self.btn_vehicle_log)
        result_layout.addWidget(vehicle_log_entry, 0)

        self.txt_schedule_debug = QPlainTextEdit(self.page_multi_result)
        self.txt_schedule_debug.setReadOnly(True)
        self.txt_schedule_debug.setMaximumBlockCount(300)
        self.txt_schedule_debug.setMaximumHeight(160)
        self.txt_schedule_debug.setPlaceholderText("排程运行日志：点击『分析当前排程』后显示前 200 条 rows 明细。")
        self.txt_schedule_debug.setStyleSheet(
            "background: #0f172a;"
            "border: 1px solid #334155;"
            "border-radius: 8px;"
            "padding: 8px;"
            "color: #e2e8f0;"
            "font-family: Menlo, Consolas, monospace;"
            "font-size: 11px;"
        )
        self.txt_schedule_debug.hide()
        result_layout.addWidget(self.txt_schedule_debug, 0)

        # ---------- Tab2：单工程组合票 ----------
        self.page_single = QWidget(self)
        page_single_layout = QVBoxLayout(self.page_single)
        page_single_layout.setContentsMargins(8, 8, 8, 8)
        page_single_layout.setSpacing(8)

        # 顶部基本信息
        row_info = QHBoxLayout()
        page_single_layout.addLayout(row_info)

        row_info.addWidget(QLabel("工程名称："))
        self.ed_sw_project = QLineEdit(self.page_single)
        self.ed_sw_project.setPlaceholderText("例如：前轴调整工位")
        self.ed_sw_project.setFixedWidth(200)
        row_info.addWidget(self.ed_sw_project)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("品番·品名："))
        self.ed_sw_part = QLineEdit(self.page_single)
        self.ed_sw_part.setPlaceholderText("例如：XXXX-XXXXX 前轮定位")
        self.ed_sw_part.setFixedWidth(220)
        row_info.addWidget(self.ed_sw_part)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("作业者："))
        self.ed_sw_worker = QLineEdit(self.page_single)
        self.ed_sw_worker.setPlaceholderText("例如：张三")
        self.ed_sw_worker.setFixedWidth(120)
        row_info.addWidget(self.ed_sw_worker)

        row_info.addSpacing(12)
        row_info.addWidget(QLabel("节拍TT(秒)："))
        self.spn_sw_takt = QSpinBox(self.page_single)
        self.spn_sw_takt.setRange(1, 9999)
        self.spn_sw_takt.setValue(118)  # 默认示例
        row_info.addWidget(self.spn_sw_takt)

        row_info.addStretch()

        # 作业手顺表（A→B 区间）
        self.tbl_sw = QTableWidget(0, 8, self.page_single)
        self.tbl_sw.setHorizontalHeaderLabels([
            "顺序", "作业名称A", "作业名称B",
            "手作业(秒)", "自动(秒)", "步行(秒)",
            "步行在前/后", "自动在前/后"
        ])
        self.tbl_sw.horizontalHeader().setStretchLastSection(True)
        self.tbl_sw.verticalHeader().setVisible(False)
        page_single_layout.addWidget(self.tbl_sw, 1)

        # 底部按钮栏（单工程组合票）
        row_btn_sw = QHBoxLayout()
        page_single_layout.addLayout(row_btn_sw)
        row_btn_sw.addStretch()

        self.btn_sw_add = QPushButton("添加作业行", self.page_single)
        self.btn_sw_del = QPushButton("删除选中行", self.page_single)
        self.btn_sw_export = QPushButton("导出标准作业组合票", self.page_single)

        row_btn_sw.addWidget(self.btn_sw_add)
        row_btn_sw.addWidget(self.btn_sw_del)
        row_btn_sw.addWidget(self.btn_sw_export)

        self.tabs.addTab(self.page_single, "单工程组合票")

        # 状态栏（用于显示导出进度 / 完成信息）
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self._on_tab_changed(self.tabs.currentIndex())
        self._update_mode_ui()

    def _connect_signals(self):
        self.btn_add_row.clicked.connect(self.add_row)
        self.btn_del_row.clicked.connect(self.del_row)
        self.btn_fill_sample.clicked.connect(self.fill_sample)
        self.cmb_launch_mode.currentIndexChanged.connect(self._update_mode_ui)
        self.btn_go_result_page.clicked.connect(lambda: self.multi_tabs.setCurrentWidget(self.page_multi_result))
        self.btn_analyze.clicked.connect(self.do_analyze)
        self.btn_export.clicked.connect(self.do_export)
        self.act_help.triggered.connect(self.show_help)
        self.btn_sim_play.clicked.connect(self._start_simulation)
        self.btn_sim_pause.clicked.connect(self._pause_simulation)
        self.btn_sim_reset.clicked.connect(self._reset_simulation)
        self.btn_vehicle_log.clicked.connect(self._show_vehicle_log_placeholder)
        self.sim_timer.timeout.connect(self._on_simulation_tick)

        # Tab 切换时，控制多工程页内步骤按钮是否可用
        self.tabs.currentChanged.connect(self._on_tab_changed)

        # 单工程组合票 Tab
        self.btn_sw_add.clicked.connect(self.add_single_row)
        self.btn_sw_del.clicked.connect(self.del_single_row)
        self.btn_sw_export.clicked.connect(self.export_single_placeholder)

    def _update_mode_ui(self):
        """根据模式切换 A/B/C 输入含义：数量模式填写数量，比例模式填写比例 + 分析时间。"""
        is_ratio = self.cmb_launch_mode.currentIndex() == 1
        if hasattr(self, "lbl_a_cars"):
            self.lbl_a_cars.setText("A比例：" if is_ratio else "A数量：")
        if hasattr(self, "lbl_b_cars"):
            self.lbl_b_cars.setText("B比例：" if is_ratio else "B数量：")
        if hasattr(self, "lbl_c_cars"):
            self.lbl_c_cars.setText("C比例：" if is_ratio else "C数量：")
        if hasattr(self, "lbl_total_cars"):
            self.lbl_total_cars.setVisible(is_ratio)
        if hasattr(self, "spn_total_cars"):
            self.spn_total_cars.setVisible(is_ratio)
        if hasattr(self, "lbl_sequence_mode"):
            self.lbl_sequence_mode.setVisible(not is_ratio)
        if hasattr(self, "cmb_seq"):
            self.cmb_seq.setVisible(not is_ratio)
        if hasattr(self, "lbl_max_run"):
            self.lbl_max_run.setVisible(not is_ratio)
        if hasattr(self, "spn_max_run"):
            self.spn_max_run.setVisible(not is_ratio)
            self.spn_max_run.setEnabled(not is_ratio)
        if hasattr(self, "params_tip"):
            if is_ratio:
                self.params_tip.setText(
                    "当前模式：按比例投车。A/B/C 填比例；分析时间填写 xx 分钟。"
                    "程序按空线起步模型，第1台从0s进入首岗位，并按目标节拍计算理论投车台数。"
                )
            else:
                self.params_tip.setText("当前模式：按数量投车。A/B/C 填数量；顺排按 A→B→C，交替混流可配合最大连续台数使用。")
    def do_analyze(self):
        try:
            project, cars, grid_step, wait_policy, defs, vehicle_counts, sequence_mode, max_consecutive, ratio_pattern, target_takt = self._collect_inputs()
            rows, max_finish = tickets.schedule(
                defs,
                cars,
                vehicle_counts,
                sequence_mode,
                max_consecutive,
                ratio_pattern,
                launch_takt=target_takt,
            )
            analysis = tickets.analyze_schedule(rows, max_finish, target_takt)
            analysis = self._apply_time_window_analysis(analysis, rows, target_takt)
            self.current_defs = list(defs or [])
            self.last_schedule_rows = rows
            self.last_analysis = analysis
            self.last_max_finish = float(max_finish or 0.0)
            if hasattr(self, "txt_schedule_debug"):
                self.txt_schedule_debug.setPlainText(self._build_schedule_debug_log(rows, limit=200))
            self.sim_time = 0.0
            self._update_sim_time_label()
            self._update_sim_total_wait_label()
            self._update_sim_view()
            self._draw_sim_scene()
            self._show_analysis_result(analysis)
            self.status.showMessage("排程分析完成", 6000)
        except Exception as e:
            import traceback
            detail = traceback.format_exc()
            print(detail)
            QMessageBox.warning(self, "分析失败", str(e))
    def _update_sim_total_wait_label(self):
        """刷新仿真控制栏中的关键判定信息。"""
        if not hasattr(self, "lbl_sim_total_wait"):
            return
        analysis = getattr(self, "last_analysis", None)
        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}
        blocking_time = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        self.lbl_sim_total_wait.setText(f"累计阻塞：{blocking_time}s")
    def _sim_speed_value(self) -> float:
        """读取仿真倍速。"""
        if not hasattr(self, "cmb_sim_speed"):
            return 1.0
        text = self.cmb_sim_speed.currentText().replace("x", "").strip()
        try:
            return max(1.0, float(text))
        except Exception:
            return 1.0

    def _update_sim_time_label(self):
        """刷新仿真时间显示。"""
        if not hasattr(self, "lbl_sim_time"):
            return
        current = float(getattr(self, "sim_time", 0.0) or 0.0)
        total = float(getattr(self, "last_max_finish", 0.0) or 0.0)
        self.lbl_sim_time.setText(f"仿真时间：{current:.1f}s / {total:.1f}s")
        if hasattr(self, "sim_progress"):
            if total > 0:
                progress_value = int(max(0.0, min(1.0, current / total)) * 1000)
            else:
                progress_value = 0
            self.sim_progress.setValue(progress_value)

    def _start_simulation(self):
        """启动仿真计时。当前阶段只推进时间，不绘制车辆。"""
        if float(getattr(self, "last_max_finish", 0.0) or 0.0) <= 0:
            QMessageBox.information(self, "提示", "请先点击『分析当前排程』生成仿真数据。")
            return
        self.sim_timer.start()

    def _pause_simulation(self):
        """暂停仿真计时。"""
        if hasattr(self, "sim_timer"):
            self.sim_timer.stop()

    def _reset_simulation(self):
        """重置仿真时间。"""
        if hasattr(self, "sim_timer"):
            self.sim_timer.stop()
        self.sim_time = 0.0
        self._update_sim_time_label()
        self._update_sim_view()
        self._draw_sim_scene()

    def _fmt_vehicle_log_value(self, value):
        if value is None:
            return ""
        try:
            num = float(value)
        except Exception:
            return str(value)
        if abs(num - round(num)) < 1e-9:
            return str(int(round(num)))
        return f"{num:.1f}"

    def _build_vehicle_log_rows(self):
        """按车辆维度整理当前 rows，仅用于只读日志弹窗。"""
        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            return [], []

        def _to_int(value, default=0):
            try:
                return int(float(value))
            except Exception:
                return default

        def _row_station(row):
            return str(row.get("step_display", row.get("station", row.get("group", ""))) or "")

        columns = [
            "车辆编号", "车型", "工序序号", "工序名称", "线别", "资源 key",
            "start", "svc_finish", "depart", "end", "dur", "block_wait", "launch_wait",
            "下一工序", "下一工序 start", "备注",
        ]

        output = []
        car_rows = self._sim_car_rows()
        sorted_car_items = sorted(
            car_rows.items(),
            key=lambda item: _to_int(item[0], 999999),
        )
        for car, segments in sorted_car_items:
            sorted_segments = sorted(
                list(segments or []),
                key=lambda row: (
                    self._sim_row_start(row),
                    _to_int(row.get("step_seq", 0)),
                ),
            )
            for idx, row in enumerate(sorted_segments):
                next_row = sorted_segments[idx + 1] if idx + 1 < len(sorted_segments) else None
                step_seq = _to_int(row.get("step_seq", ""), 0)
                next_seq = _to_int(next_row.get("step_seq", ""), 0) if next_row else 0
                note = ""
                if next_row and step_seq > 0 and next_seq > step_seq + 1:
                    skipped = ", ".join(str(seq) for seq in range(step_seq + 1, next_seq))
                    note = f"跳过中间岗位：{skipped}"

                output.append([
                    str(car),
                    str(row.get("car_type", row.get("duration_source", row.get("vehicle_type", ""))) or ""),
                    str(row.get("step_seq", "")),
                    _row_station(row),
                    str(row.get("line_no", row.get("line", "")) or ""),
                    str(row.get("resource_key", "") or ""),
                    self._fmt_vehicle_log_value(row.get("start", row.get("start_time", ""))),
                    self._fmt_vehicle_log_value(row.get("svc_finish", row.get("finish", ""))),
                    self._fmt_vehicle_log_value(row.get("depart", row.get("end", ""))),
                    self._fmt_vehicle_log_value(row.get("end", "")),
                    self._fmt_vehicle_log_value(row.get("dur", row.get("duration", ""))),
                    self._fmt_vehicle_log_value(row.get("block_wait", 0.0)),
                    self._fmt_vehicle_log_value(row.get("launch_wait", 0.0)),
                    _row_station(next_row) if next_row else "无",
                    self._fmt_vehicle_log_value(next_row.get("start", next_row.get("start_time", ""))) if next_row else "",
                    note,
                ])

        return columns, output

    def _copy_vehicle_log_to_clipboard(self, columns, rows):
        lines = ["\t".join(columns)]
        lines.extend("\t".join(str(value) for value in row) for row in rows)
        QApplication.clipboard().setText("\n".join(lines))
        self.status.showMessage("车辆日志已复制到剪贴板", 3000)

    def _show_vehicle_log_placeholder(self):
        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            QMessageBox.information(self, "车辆明细 / 调试日志", "请先点击“分析当前排程”生成车辆日志。")
            return

        columns, log_rows = self._build_vehicle_log_rows()
        vehicle_count = len(self._sim_car_rows())

        dialog = QDialog(self)
        dialog.setWindowTitle("车辆明细 / 调试日志")
        dialog.resize(1000, 650)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        info = QLabel(
            f"只读排程明细：rows 共 {len(rows)} 条，车辆 {vehicle_count} 台。"
            "下一工序基于同一车辆 rows 顺序计算；用于排查 depart / block_wait / 跳过岗位。"
        )
        info.setWordWrap(True)
        info.setStyleSheet("color:#475569;font-size:12px;")
        layout.addWidget(info)

        table = QTableWidget(len(log_rows), len(columns), dialog)
        table.setHorizontalHeaderLabels(columns)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)

        for r, row_values in enumerate(log_rows):
            for c, value in enumerate(row_values):
                item = QTableWidgetItem(str(value))
                if c in (0, 2, 6, 7, 8, 9, 10, 11, 12, 14):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                table.setItem(r, c, item)

        table.horizontalHeader().setStretchLastSection(True)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.resizeColumnsToContents()
        table.setSortingEnabled(True)
        layout.addWidget(table, 1)

        button_row = QHBoxLayout()
        button_row.addStretch()
        btn_copy = QPushButton("复制全部日志", dialog)
        btn_close = QPushButton("关闭", dialog)
        button_row.addWidget(btn_copy)
        button_row.addWidget(btn_close)
        layout.addLayout(button_row)

        btn_copy.clicked.connect(lambda: self._copy_vehicle_log_to_clipboard(columns, log_rows))
        btn_close.clicked.connect(dialog.accept)

        dialog.exec()

    def _on_simulation_tick(self):
        """仿真计时推进。后续车辆绘制会基于 sim_time 刷新画面。"""
        total = float(getattr(self, "last_max_finish", 0.0) or 0.0)
        if total <= 0:
            self._pause_simulation()
            return
        self.sim_time = min(total, float(getattr(self, "sim_time", 0.0) or 0.0) + 0.1 * self._sim_speed_value())
        self._update_sim_time_label()
        self._update_sim_view()
        self._draw_sim_scene()
        if self.sim_time >= total:
            self._pause_simulation()


    def _sim_row_value(self, row: dict, *keys, default=None):
        """从排程行中兼容读取字段。"""
        for key in keys:
            if key in row and row.get(key) is not None:
                return row.get(key)
        return default

    def _sim_row_start(self, row: dict) -> float:
        value = self._sim_row_value(row, "start", "start_time", "begin", "in", "in_time", default=0.0)
        try:
            return float(value or 0.0)
        except Exception:
            return 0.0

    def _sim_row_end(self, row: dict) -> float:
        value = self._sim_row_value(row, "end", "finish", "finish_time", "out", "out_time", default=None)
        if value is not None:
            try:
                return float(value or 0.0)
            except Exception:
                pass
        start = self._sim_row_start(row)
        dur = self._sim_row_value(row, "dur", "duration", "process", "process_time", default=0.0)
        try:
            return start + float(dur or 0.0)
        except Exception:
            return start

    def _sim_row_service_finish(self, row: dict) -> float:
        """读取当前工程加工完成时间，优先使用 svc_finish。"""
        value = self._sim_row_value(
            row,
            "svc_finish",
            "finish",
            "finish_time",
            "end",
            default=None,
        )
        if value is not None:
            try:
                return float(value or 0.0)
            except Exception:
                pass
        return self._sim_row_end(row)

    def _sim_row_depart(self, row: dict) -> float:
        """读取车辆可离开当前工程的时间，优先使用 depart。"""
        value = self._sim_row_value(
            row,
            "depart",
            "end",
            "out",
            "out_time",
            "svc_finish",
            "finish",
            default=None,
        )
        if value is not None:
            try:
                return float(value or 0.0)
            except Exception:
                pass
        return max(self._sim_row_service_finish(row), self._sim_row_end(row))

    def _sim_row_wait_end(self, row: dict, next_row: dict | None = None) -> float:
        """
        等待阶段的显示终点。

        规则：
        - 优先停留在当前工程所在位置；
        - depart/end 只作为“仍停留在当前工程”的参考；
        - 如果下一工程尚未开始，current < next_start 时不允许提前跳到 next_row。
        """
        wait_start = self._sim_row_service_finish(row)
        wait_end = max(wait_start, self._sim_row_depart(row))
        if next_row is not None:
            wait_end = max(wait_end, self._sim_row_start(next_row))
        return wait_end

    def _sim_row_station(self, row: dict) -> str:
        return str(self._sim_row_value(row, "step_display", "station", "display", "name", "group", default="岗位") or "岗位")

    def _sim_row_car_label(self, row: dict) -> str:
        car = self._sim_row_value(row, "car", "car_no", "car_index", "idx", default="?")
        car_type = str(self._sim_row_value(row, "car_type", "type", "vehicle_type", default="") or "")
        if car_type:
            return f"Car#{car}({car_type})"
        return f"Car#{car}"
    
    def _sim_row_run_mode(self, row: dict) -> str:
        """兼容读取岗位运行方式。"""
        return str(self._sim_row_value(row, "run_mode", "mode", default="") or "")

    def _sim_row_line_no(self, row: dict) -> str:
        """兼容读取排程行线别，用于 v2-4 线别验证。"""
        return str(self._sim_row_value(row, "line_no", "line", "line_scope", default="") or "")

    def _sim_row_resource_key(self, row: dict) -> str:
        """兼容读取排程行资源 key。"""
        return str(self._sim_row_value(row, "resource_key", default="") or "")

    def _sim_row_block_wait(self, row: dict) -> float:
        """读取当前工程后方阻塞等待时间，仅用于 UI 保守显示。"""
        value = self._sim_row_value(row, "block_wait", default=0.0)
        try:
            return float(value or 0.0)
        except Exception:
            return 0.0

    def _sim_wait_is_blocking(self, row: dict, next_row: dict | None, occupied_resources: set[str]) -> bool:
        """
        UI 安全模式下，默认不把等待解释为“阻塞”。

        说明：
        - 当前项目的稳定诉求是保留新画布结构，但收缩显示层解释；
        - finish -> next_start 之间的时间差，不应由 UI 轻易推断成“被谁阻塞”；
        - process_over_takt_root_text 属于模型结果说明，也不用于实时阻塞判定。

        因此这里保守返回 False，等待文案统一走“等待Xs”。
        未来如果 rows 中新增了明确、可靠的实时阻塞字段，再单独恢复更细判断。
        """
        return False

    def _sim_wait_label(self, row: dict, next_row: dict | None, current: float, wait_end: float, occupied_resources: set[str]) -> str:
        """统一生成等待文案；安全模式下默认只显示等待时长。"""
        wait_remain = max(0.0, wait_end - current)
        return f"等待{wait_remain:.1f}s"

    def _build_schedule_debug_log(self, rows, limit: int = 200) -> str:
        """临时排程运行日志，用于 v2 排程模型验证；后续稳定后可隐藏或删除。"""
        if not rows:
            return "暂无排程 rows。"

        lines = [
            "排程运行日志（临时验证用）",
            "字段：Car | 岗位 | 线别 | 资源 | 理论投车 | 开始 | 完成 | 离开 | 工时 | 等待",
            "-" * 120,
        ]

        def _fmt(value):
            try:
                return f"{float(value or 0.0):.1f}"
            except Exception:
                return str(value)

        sorted_rows = sorted(
            list(rows or []),
            key=lambda r: (
                int(r.get("car", r.get("car_no", r.get("car_index", 0))) or 0),
                float(r.get("start", r.get("start_time", 0.0)) or 0.0),
                int(r.get("step_seq", 0) or 0),
            ),
        )

        for idx, row in enumerate(sorted_rows[:limit], start=1):
            car = row.get("car", row.get("car_no", row.get("car_index", "?")))
            car_type = str(row.get("car_type", row.get("duration_source", row.get("vehicle_type", ""))) or "")
            station = str(row.get("step_display", row.get("station", row.get("group", "岗位"))) or "岗位")
            line_no = str(row.get("line_no", row.get("line", "")) or "")
            resource_key = str(row.get("resource_key", "") or "")
            theory_launch = row.get("theory_launch_time", "")
            start = row.get("start", row.get("start_time", 0.0))
            svc_finish = row.get("svc_finish", row.get("finish", row.get("end", 0.0)))
            depart = row.get("depart", row.get("end", row.get("svc_finish", 0.0)))
            dur = row.get("dur", row.get("duration", 0.0))
            wait = float(row.get("block_wait", 0.0) or 0.0) + float(row.get("launch_wait", 0.0) or 0.0)

            lines.append(
                f"{idx:03d}. Car#{car}({car_type}) | {station} | "
                f"线别 {line_no or '—'} | 资源 {resource_key or '—'} | "
                f"理论投车 {_fmt(theory_launch)}s | 开始 {_fmt(start)}s | "
                f"完成 {_fmt(svc_finish)}s | 离开 {_fmt(depart)}s | "
                f"工时 {_fmt(dur)}s | 等待 {_fmt(wait)}s"
            )

        if len(sorted_rows) > limit:
            lines.append(f"……仅显示前 {limit} 条，共 {len(sorted_rows)} 条。")

        return "\n".join(lines)
    
    def _sim_car_key(self, row: dict):
        """按车辆编号聚合排程段。"""
        return self._sim_row_value(row, "car", "car_no", "car_index", "idx", default="?")
    
    def _sim_car_rows(self):
        """将排程行按车辆聚合，并按开始时间排序。"""
        grouped = {}
        for row in getattr(self, "last_schedule_rows", []) or []:
            key = self._sim_car_key(row)
            grouped.setdefault(key, []).append(row)
        for key in grouped:
            grouped[key].sort(key=lambda r: self._sim_row_start(r))
        return grouped

    def _sim_station_names(self):
        """优先按岗位矩阵 seq 顺序提取岗位名。"""
        defs = getattr(self, "current_defs", None) or []
        if defs:
            ordered = []
            seen = set()
            sortable_defs = []
            for idx, item in enumerate(defs):
                try:
                    seq_value = int(item.get("seq", idx + 1) or (idx + 1))
                except Exception:
                    seq_value = idx + 1
                name = str(item.get("display") or item.get("group") or "岗位").strip() or "岗位"
                sortable_defs.append((seq_value, idx, name))
            sortable_defs.sort(key=lambda x: (x[0], x[1]))
            for _, _, name in sortable_defs:
                if name not in seen:
                    ordered.append(name)
                    seen.add(name)
            if ordered:
                return ordered

        names = []
        seen = set()
        for row in getattr(self, "last_schedule_rows", []) or []:
            name = self._sim_row_station(row)
            if name not in seen:
                names.append(name)
                seen.add(name)
        return names

    def _sim_station_defs(self):
        """返回动画展示使用的岗位顺序与标签，优先使用 current_defs 的 seq。"""
        defs = getattr(self, "current_defs", None) or []
        if defs:
            ordered = []
            seen = set()
            sortable_defs = []
            for idx, item in enumerate(defs):
                try:
                    seq_value = int(item.get("seq", idx + 1) or (idx + 1))
                except Exception:
                    seq_value = idx + 1
                name = str(item.get("display") or item.get("group") or "岗位").strip() or "岗位"
                sortable_defs.append((seq_value, idx, name, item))
            sortable_defs.sort(key=lambda x: (x[0], x[1]))
            for seq_value, _, name, source in sortable_defs:
                if name not in seen:
                    ordered.append({
                        "seq": seq_value,
                        "name": name,
                        "device_count": source.get("device_count"),
                        "line_scope": source.get("line_scope"),
                        "run_mode": source.get("run_mode"),
                    })
                    seen.add(name)
            if ordered:
                return ordered

        names = self._sim_station_names()
        return [{"seq": idx + 1, "name": name} for idx, name in enumerate(names)]

    def _update_sim_view(self):
        """刷新简易仿真画面。当前阶段只展示当前加工中的车辆摘要。"""
        if not hasattr(self, "lbl_sim_view"):
            return

        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            self.lbl_sim_view.setText("仿真画面：请先点击『分析当前排程』。")
            return

        current = float(getattr(self, "sim_time", 0.0) or 0.0)

        active_rows = []
        for row in rows:
            start = self._sim_row_start(row)
            svc_finish = self._sim_row_service_finish(row)
            if start <= current < svc_finish:
                active_rows.append((start, svc_finish, row))

        active_rows.sort(key=lambda x: (self._sim_row_station(x[2]), x[0]))
        waiting_rows = []
        car_rows = self._sim_car_rows()
        for _, car_segments in car_rows.items():
            if not car_segments:
                continue
            for idx, seg in enumerate(car_segments):
                next_seg = car_segments[idx + 1] if idx + 1 < len(car_segments) else None
                wait_start = self._sim_row_service_finish(seg)
                wait_end = self._sim_row_wait_end(seg, next_seg)
                if wait_start <= current < wait_end:
                    waiting_rows.append((wait_start, wait_end, seg, next_seg))
                    break

        waiting_rows.sort(key=lambda x: (self._sim_row_station(x[2]), x[0]))

        def _escape_lines(lines):
            return [escape(str(line)) for line in lines if str(line).strip()]

        def _compress_lines(lines, max_lines):
            cleaned = [str(line).strip() for line in lines if str(line).strip()]
            if not cleaned:
                return ["无"]
            if len(cleaned) <= max_lines:
                return cleaned
            kept = cleaned[:max_lines]
            overflow = len(cleaned) - max_lines
            kept[-1] = f"{kept[-1]} / ……还有 {overflow} 台"
            return kept

        def _slot_html(lines):
            escaped_lines = _escape_lines(lines)
            while len(escaped_lines) < 2:
                escaped_lines.append("&#12288;")
            return "".join(
                f"<div style='margin:0;padding:0;line-height:1.28;'>{line}</div>"
                for line in escaped_lines[:2]
            )

        active_lines = []
        if active_rows:
            max_active_rows = 3
            for start, end, row in active_rows[:max_active_rows]:
                car_label = self._sim_row_car_label(row)
                station = self._sim_row_station(row)
                remain = max(0.0, end - current)
                line_no = self._sim_row_line_no(row)
                detail_parts = [f"{car_label} @ {station}"]
                if line_no:
                    detail_parts.append(str(line_no))
                if remain > 0:
                    detail_parts.append(f"剩余{remain:.1f}s")
                active_lines.append("｜".join(detail_parts))
            if len(active_rows) > max_active_rows:
                active_lines.append(f"……还有 {len(active_rows) - max_active_rows} 台")
        else:
            active_lines.append("无")
        active_lines = _compress_lines(active_lines, 2)

        waiting_lines = []
        if waiting_rows:
            max_wait_rows = 2
            for wait_start, wait_end, row, next_row in waiting_rows[:max_wait_rows]:
                car_label = self._sim_row_car_label(row)
                station = self._sim_row_station(row)
                line_no = self._sim_row_line_no(row)
                detail_parts = [f"{car_label} @ {station}"]
                if line_no:
                    detail_parts.append(str(line_no))
                wait_text = self._sim_wait_label(row, next_row, current, wait_end, set())
                if wait_text:
                    detail_parts.append(wait_text)
                waiting_lines.append("｜".join(detail_parts))
            if len(waiting_rows) > max_wait_rows:
                waiting_lines.append(f"……还有 {len(waiting_rows) - max_wait_rows} 台")
        else:
            waiting_lines.append("无")
        waiting_lines = _compress_lines(waiting_lines, 2)

        analysis = getattr(self, "last_analysis", None)
        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}
        model_hint_text = str(
            summary.get("process_over_takt_root_text")
            or summary.get("process_root_text")
            or "无"
        ).strip() or "无"
        if model_hint_text != "无":
            normalized_blocking = (
                model_hint_text.replace("、", "，")
                .replace(",", "，")
                .replace("；", "，")
                .replace(";", "，")
            )
            blocking_parts = [
                part.strip() for part in normalized_blocking.split("，") if part.strip()
            ]
            if len(blocking_parts) > 2:
                model_hint_text = f"超节拍工程：{'，'.join(blocking_parts[:2])} 等"
            elif blocking_parts:
                model_hint_text = f"超节拍工程：{'，'.join(blocking_parts)}"
        else:
            model_hint_text = "详见模型结果"
        blocking_lines = _compress_lines(model_hint_text.splitlines() or [model_hint_text], 2)

        active_html = _slot_html(active_lines)
        waiting_html = _slot_html(waiting_lines)
        blocking_html = _slot_html(blocking_lines)

        html = (
            "<table width='100%' cellspacing='0' cellpadding='0'>"
            "<tr>"
            "<td width='33%' style='width:33%;vertical-align:top;padding-right:8px;'>"
            "<div style='font-size:12px;font-weight:700;color:#334155;margin-bottom:2px;'>加工中车辆</div>"
            f"<div style='font-size:12px;line-height:1.24;color:#334155;'>{active_html}</div>"
            "</td>"
            "<td width='34%' style='width:34%;vertical-align:top;padding:0 8px;border-left:1px solid #e2e8f0;'>"
            "<div style='font-size:12px;font-weight:700;color:#334155;margin-bottom:2px;'>等待中</div>"
            f"<div style='font-size:12px;line-height:1.24;color:#334155;'>{waiting_html}</div>"
            "</td>"
            "<td width='33%' style='width:33%;vertical-align:top;padding-left:8px;border-left:1px solid #e2e8f0;'>"
            "<div style='font-size:12px;font-weight:700;color:#334155;margin-bottom:2px;'>模型提示</div>"
            f"<div style='font-size:12px;line-height:1.24;color:#334155;'>{blocking_html}</div>"
            "</td>"
            "</tr>"
            "</table>"
        )
        self.lbl_sim_view.setText(html)

    # ------------- 多车组合票：动作 ------------- #
    def add_row(self):
        r = self.tbl.rowCount()
        self.tbl.insertRow(r)

        # 序号
        self.tbl.setItem(r, 0, QTableWidgetItem(str(r + 1)))
        # 工程名称
        self.tbl.setItem(r, 1, QTableWidgetItem(""))

        # 设备数量：v2 主线字段。默认 2，代表双线双设备。
        device_count_cb = QComboBox(self.tbl)
        device_count_cb.addItems(["1", "2"])
        device_count_cb.setCurrentText("2")
        self.tbl.setCellWidget(r, 2, device_count_cb)

        # 所属线别：v2 主线字段。
        line_scope_cb = QComboBox(self.tbl)
        line_scope_cb.addItems(["1号线", "2号线", "双线", "双线共用"])
        line_scope_cb.setCurrentText("双线")
        self.tbl.setCellWidget(r, 3, line_scope_cb)

        # 岗位设备
        self.tbl.setItem(r, 4, QTableWidgetItem(""))
        # A / B / C 工时
        self.tbl.setItem(r, 5, QTableWidgetItem(""))
        self.tbl.setItem(r, 6, QTableWidgetItem(""))
        self.tbl.setItem(r, 7, QTableWidgetItem(""))

        def _sync_line_scope():
            if device_count_cb.currentText() == "2":
                line_scope_cb.setCurrentText("双线")
                line_scope_cb.setEnabled(False)
            else:
                line_scope_cb.setEnabled(True)
                if line_scope_cb.currentText() == "双线":
                    line_scope_cb.setCurrentText("1号线")

        device_count_cb.currentTextChanged.connect(_sync_line_scope)
        _sync_line_scope()

    def _choose_color(self, row: int):
        dlg_col = QColorDialog.getColor(parent=self)
        if dlg_col.isValid():
            hex_code = dlg_col.name()
            btn = self.tbl.cellWidget(row, self.COL_C_TIME)
            btn.setStyleSheet(f"background:{hex_code};")
            self.tbl.item(row, self.COL_C_TIME).setData(Qt.UserRole, hex_code)

    def del_row(self):
        r = self.tbl.currentRow()
        if r >= 0:
            self.tbl.removeRow(r)

    # -------- 单人标准作业组合票：行操作 --------
    def add_single_row(self):
        """在单人作业手顺表中新增一行"""
        if not hasattr(self, "tbl_sw"):
            return

        current_rows = self.tbl_sw.rowCount()
        max_steps = getattr(self, "MAX_SINGLE_STEPS", 23)

        if current_rows >= max_steps:
            QMessageBox.warning(
                self,
                "已到模板上限",
                f"当前单人标准作业组合票模板最多支持 {max_steps} 行。\n"
                f"你现在已经添加了 {current_rows} 行，不能再继续新增。\n\n"
                "请合并部分区间或拆分为多张组合票后再导出。"
            )
            return

        r = current_rows
        self.tbl_sw.insertRow(r)
        # 顺序默认递增（组合票行号）
        self.tbl_sw.setItem(r, 0, QTableWidgetItem(str(r + 1)))
        # 作业名称A / B 先留空，让你填写
        self.tbl_sw.setItem(r, 1, QTableWidgetItem(""))
        self.tbl_sw.setItem(r, 2, QTableWidgetItem(""))
        # 手作业 / 自动 / 步行，默认 0
        self.tbl_sw.setItem(r, 3, QTableWidgetItem("0"))
        self.tbl_sw.setItem(r, 4, QTableWidgetItem("0"))
        self.tbl_sw.setItem(r, 5, QTableWidgetItem("0"))
        # 步行位置：默认“后置”
        pos_cb = QComboBox(self.tbl_sw)
        pos_cb.addItem("后置", userData="after")
        pos_cb.addItem("前置", userData="before")
        self.tbl_sw.setCellWidget(r, 6, pos_cb)
        # 自动在前/后（默认后置）
        auto_cb = QComboBox(self.tbl_sw)
        auto_cb.addItem("后置", userData="after")
        auto_cb.addItem("前置", userData="before")
        self.tbl_sw.setCellWidget(r, 7, auto_cb)

    def del_single_row(self):
        """删除单人作业手顺表中的选中行"""
        if not hasattr(self, "tbl_sw"):
            return
        r = self.tbl_sw.currentRow()
        if r >= 0:
            self.tbl_sw.removeRow(r)
        # 重写顺序列，让它保持 1,2,3,...
        for i in range(self.tbl_sw.rowCount()):
            item = self.tbl_sw.item(i, 0)
            if item is None:
                item = QTableWidgetItem()
                self.tbl_sw.setItem(i, 0, item)
            item.setText(str(i + 1))

    # -------- 单人标准作业组合票：数据收集 --------
    def _collect_single_inputs(self):
        """
        从单人作业手顺 Tab 中读取数据，并计算时间汇总。
        返回：
          project, part, worker, takt_sec, steps, totals
        其中：
          steps: [{seq, name, name_a, name_b, manual, auto, walk, walk_pos, auto_pos, duration, start, end}, ...]
          totals: {"manual": x, "auto": y, "walk": z, "total": t}
        """
        if not hasattr(self, "tbl_sw"):
            raise ValueError("单人作业手顺表尚未初始化")

        project = (self.ed_sw_project.text().strip() or "工程")
        part = self.ed_sw_part.text().strip()
        worker = self.ed_sw_worker.text().strip()
        takt_sec = int(self.spn_sw_takt.value())

        steps = []
        cur_time = 0.0
        total_manual = 0.0
        total_auto = 0.0
        total_walk = 0.0

        for r in range(self.tbl_sw.rowCount()):
            # 作业名称 A / B
            name_a_item = self.tbl_sw.item(r, 1)
            name_b_item = self.tbl_sw.item(r, 2)
            name_a = name_a_item.text().strip() if name_a_item else ""
            name_b = name_b_item.text().strip() if name_b_item else ""

            if not name_a and not name_b:
                # 两个都没填，当作空行，跳过
                continue

            # 导出时使用的显示名（A→B / 单独一个）
            if name_a and name_b:
                name = f"{name_a} → {name_b}"
            else:
                name = name_a or name_b

            def _get_time(col_idx: int) -> float:
                item = self.tbl_sw.item(r, col_idx)
                txt = item.text().strip() if item else ""
                if not txt:
                    return 0.0
                try:
                    return float(txt)
                except Exception:
                    raise ValueError(f"第 {r + 1} 行时间列（第 {col_idx + 1} 列）不是有效数字：{txt}")

            # 手作业 / 自动 / 步行时间列：3, 4, 5
            manual = _get_time(3)
            auto = _get_time(4)
            walk = _get_time(5)

            # 步行位置：前置/后置（默认后置）
            walk_pos = "after"
            pos_widget = self.tbl_sw.cellWidget(r, 6)
            if isinstance(pos_widget, QComboBox):
                walk_pos_data = pos_widget.currentData()
                if walk_pos_data in ("before", "after"):
                    walk_pos = walk_pos_data

            # 自动在前/后（默认后置）
            auto_pos = "after"
            auto_widget = self.tbl_sw.cellWidget(r, 7)
            if isinstance(auto_widget, QComboBox):
                auto_pos_data = auto_widget.currentData()
                if auto_pos_data in ("before", "after"):
                    auto_pos = auto_pos_data

            duration = manual + auto + walk
            if duration <= 0:
                raise ValueError(f"第 {r + 1} 行『{name}』的时间合计为 0，请填写手作业/自动/步行时间。")

            start = cur_time
            end = cur_time + duration
            cur_time = end

            total_manual += manual
            total_auto += auto
            total_walk += walk

            # 顺序列（如果用户改过，我们尽量读取）
            seq_item = self.tbl_sw.item(r, 0)
            try:
                seq = int(seq_item.text()) if seq_item and seq_item.text().strip() else len(steps) + 1
            except Exception:
                seq = len(steps) + 1

            steps.append({
                "seq": seq,
                "name": name,       # A→B 组合显示名（保留）
                "name_a": name_a,   # 原始作业名称A
                "name_b": name_b,   # 原始作业名称B
                "manual": manual,
                "auto": auto,
                "walk": walk,
                "walk_pos": walk_pos,  # 步行在前/后
                "auto_pos": auto_pos,  # 自动在前/后
                "duration": duration,
                "start": start,
                "end": end,
            })

        # 行数上限检查：防止超过模板预留的行数
        if len(steps) > self.MAX_SINGLE_STEPS:
            raise ValueError(
                f"当前单人标准作业组合票共有 {len(steps)} 行，已超过模板最多支持的 {self.MAX_SINGLE_STEPS} 行。\n"
                "请合并部分区间或拆分为多张组合票后再导出。"
            )

        if not steps:
            raise ValueError("请至少填写一行有效的作业（需有作业名称和时间）。")

        totals = {
            "manual": total_manual,
            "auto": total_auto,
            "walk": total_walk,
            "total": total_manual + total_auto + total_walk,
        }
        return project, part, worker, takt_sec, steps, totals

    # -------- 单人标准作业组合票：写入模板 --------
    def _export_single_to_excel(self, path, project, part, worker, takt_sec, steps, totals):
        """
        根据单人作业手顺（A→B 区间）将数据写入《组合票标准版.xlsx》模板：
        - 模板文件需放在与本文件同一目录下，文件名：组合票标准版.xlsx
        - 仅填充左侧步骤表区域和基本信息，不修改模板中的其他格式/图表
        """
        # 定位模板文件：与本 .py 同目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(base_dir, "组合票标准版.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"未找到模板文件：{template_path}")

        wb = load_workbook(template_path)
        try:
            ws = wb["④标准作业组合票"]
        except KeyError:
            ws = wb.active

        def _set_value(coord, value):
            """安全写入单元格：若目标是合并单元格，从其合并区域左上角写入"""
            cell = ws[coord]
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        ws.cell(row=mr.min_row, column=mr.min_col).value = value
                        break
            else:
                cell.value = value

        def _set_fill(row, col, fill):
            """安全设置单元格填充：若目标是合并单元格，则写到其合并区域左上角"""
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        ws.cell(row=mr.min_row, column=mr.min_col).fill = fill
                        break
            else:
                cell.fill = fill

        def _set_border(row, col, border: Border):
            """
            安全设置单元格边框：若目标是合并单元格，则写到其合并区域左上角；
            与已有边框合并（只改指定方向的线型）。
            """
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break

            old = cell.border or Border()

            def merge_side(new_side, old_side):
                if getattr(new_side, "style", None):
                    return new_side
                return old_side

            cell.border = Border(
                left=merge_side(border.left, old.left),
                right=merge_side(border.right, old.right),
                top=merge_side(border.top, old.top),
                bottom=merge_side(border.bottom, old.bottom),
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        def _clear_top_border(row, col):
            """
            清除单元格的上边框（保留其余边框），合并单元格时操作左上角单元格。
            """
            cell = ws.cell(row=row, column=col)
            # Handle merged cells: always operate on effective top-left cell
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break
            old = cell.border or Border()
            cell.border = Border(
                left=old.left,
                right=old.right,
                top=Side(style=None),
                bottom=old.bottom,
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        def _clear_left_border(row, col):
            """
            清除单元格的左边框（保留其余边框）；合并单元格时操作左上角单元格。
            """
            cell = ws.cell(row=row, column=col)
            # Handle merged cells: always operate on effective top-left cell
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break
            old = cell.border or Border()
            cell.border = Border(
                left=Side(style=None),
                right=old.right,
                top=old.top,
                bottom=old.bottom,
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        def _clear_right_border(row, col):
            """
            清除单元格的右边框（保留其余边框）；合并单元格时操作左上角单元格。
            """
            cell = ws.cell(row=row, column=col)
            # Handle merged cells: always operate on effective top-left cell
            if isinstance(cell, MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = ws.cell(row=mr.min_row, column=mr.min_col)
                        break
            old = cell.border or Border()
            cell.border = Border(
                left=old.left,
                right=Side(style=None),
                top=old.top,
                bottom=old.bottom,
                diagonal=old.diagonal,
                diagonal_direction=old.diagonal_direction,
                outline=old.outline,
                vertical=old.vertical,
                horizontal=old.horizontal,
            )

        # 1) 清空左侧原有数据区域
        start_row = 9
        row_span = 3
        max_steps = getattr(self, "MAX_SINGLE_STEPS", 23)
        end_row = start_row + max_steps * row_span - 1
        for row in range(start_row, end_row + 1):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None

        # 清空右侧时间轴区域填充（F列开始，按总时间估算范围）
        time_start_col = 6  # F列
        max_time = int(round(totals.get("total", 0))) if isinstance(totals, dict) else 0
        if max_time < 0:
            max_time = 0
        time_end_col = time_start_col + max_time + 5
        for row in range(start_row, end_row + 1):
            for col in range(time_start_col, time_end_col + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = PatternFill()

          # 2) 写入步骤：每步占 3 行（A9:A11, A12:A14, ...）
        row_span = 3
        time_start_col = 6  # F列，时间轴起始列
        time_fill = PatternFill(fill_type="solid", fgColor="000000")
        # 自动：加粗虚线（仅画上边框，不填充）
        auto_side = Side(style="mediumDashed", color="000000")
        h_auto_border = Border(top=auto_side)

        # segments：按“步骤”记录每一行的黑条起止（即手作业+自动的时间段）
        segments = []  # [{"mid_row": int, "bar_start": int, "bar_end": int}, ...]

        for idx, s in enumerate(steps):
            base_row = start_row + idx * row_span

            # 序号
            ws.cell(row=base_row, column=1).value = s["seq"]

            # 作业名称 A/B：B 列两行
            name_a = s.get("name_a") or s.get("name") or ""
            name_b = s.get("name_b") or ""
            ws.cell(row=base_row, column=2).value = name_a
            if name_b:
                ws.cell(row=base_row + 2, column=2).value = name_b

            # 时间数值（C~E）
            ws.cell(row=base_row, column=3).value = s["manual"]
            ws.cell(row=base_row, column=4).value = s["auto"]
            ws.cell(row=base_row, column=5).value = s["walk"]

            # ===== 时间轴绘制（手作业=黑填充；自动=加粗虚线；步行仅用折线表示） =====
            start_sec = int(round(s["start"]))
            manual = float(s["manual"])
            auto = float(s["auto"])
            walk = float(s["walk"])
            walk_pos = s.get("walk_pos", "after")
            auto_pos = s.get("auto_pos", "after")

            # 起点：若步行在前，整体右移
            if walk_pos == "before":
                bar_start_sec = int(round(start_sec + walk))
            else:
                bar_start_sec = start_sec

            mid_row = base_row + 1

            # 决定绘制顺序：自动在前/后
            draw_seq = []
            if auto_pos == "before":
                if auto > 0:
                    draw_seq.append(("auto", auto))
                if manual > 0:
                    draw_seq.append(("manual", manual))
            else:
                if manual > 0:
                    draw_seq.append(("manual", manual))
                if auto > 0:
                    draw_seq.append(("auto", auto))

            seg_start = bar_start_sec
            for kind, length in draw_seq:
                seg_end = int(round(seg_start + length))
                if seg_end > seg_start:
                    for sec in range(seg_start, seg_end):
                        col = time_start_col + sec
                        if kind == "manual":
                            _set_fill(mid_row, col, time_fill)        # 手作业：黑色填充
                        else:
                            _set_border(mid_row, col, h_auto_border)  # 自动：加粗虚线（上边框）
                seg_start = seg_end

            # 记录该步的整体开始/结束（不包含步行在后）
            bar_end_sec = seg_start
            segments.append(
                {
                    "mid_row": mid_row,
                    "bar_start": bar_start_sec,
                    "bar_end": bar_end_sec,
                }
            )

        # 2.5) 相邻「步骤」之间画连接线：
        #      - 有间隔：步行 → 实折线，从黑条末端右边一格开始，先竖后横
        #      - 无间隔：直接接续 → 加粗实直线
        if len(segments) >= 2:
            solid_side = Side(style="medium", color="000000")   # 加粗实线
            walk_side  = Side(style="medium", color="000000")   # 步行：加粗实线

            h_walk_border = Border(top=walk_side)               # 步行横线
            v_walk_left   = Border(left=walk_side)              # 竖线（当前列左边）
            v_walk_right  = Border(right=walk_side)             # 竖线镜像（前一列右边）
            v_solid_right_border = Border(right=solid_side)     # 无间隔直连竖线（边界线上）

            for i in range(len(segments) - 1):
                curr = segments[i]
                nxt = segments[i + 1]

                mid_row_curr = curr["mid_row"]
                mid_row_nxt = nxt["mid_row"]
                bar_end_curr = curr["bar_end"]
                bar_start_nxt = nxt["bar_start"]

                # 注意：bar_end / bar_start 是“时间（秒）”，还没加上 F 列偏移
                if bar_start_nxt > bar_end_curr:
                    # 有间隔：步行 → 实折线
                    # 连接策略：
                    #   - 竖线画在“上一列的右边界”，并且只画到下一段所在行的上一行（不进入下一段单元格）
                    #   - 横线从拐点所在列开始，沿下一段所在行的上边框一直画到下一段条形左侧
                    first_blank_col = time_start_col + bar_end_curr        # 上一段末尾右侧的第一格
                    next_bar_first_col = time_start_col + bar_start_nxt    # 下一段条形开始列

                    # 1) 竖线：用上一列（first_blank_col - 1）的『右边界』画，恰好停在下一段顶边
                    grid_col_for_right_edge = first_blank_col - 1
                    row_vert_start = mid_row_curr + 1
                    row_vert_end_exclusive = mid_row_nxt  # 不包含下一段所在行，避免出现“下垂尾巴”
                    if grid_col_for_right_edge >= time_start_col and row_vert_start < row_vert_end_exclusive:
                        for row in range(row_vert_start, row_vert_end_exclusive):
                            _set_border(row, grid_col_for_right_edge, v_walk_right)
                    # 保底清理下一行该列的右边界，避免‘下垂尾巴’
                    _clear_right_border(mid_row_nxt, grid_col_for_right_edge)

                    # 2) 横线：从拐点所在列开始（不跳空），一直到下一段左侧列
                    start_h_col = first_blank_col  # 不留缺口
                    if start_h_col < next_bar_first_col:
                        for col in range(start_h_col, next_bar_first_col):
                            _set_border(mid_row_nxt, col, h_walk_border)
                else:
                    # 无间隔：在上一段最后一秒所在列的“右边界”连线，
                    # 竖线落在列缝而不是下一段条形内部，且不覆盖下一段所在行
                    boundary_col = time_start_col + bar_end_curr
                    row_top = min(mid_row_curr, mid_row_nxt)
                    row_bottom = max(mid_row_curr, mid_row_nxt) - 1
                    if row_top <= row_bottom:
                        for row in range(row_top, row_bottom + 1):
                            _set_border(row, boundary_col - 1, v_solid_right_border)
                    # 保底清理下一行该列的右边界，避免‘下垂尾巴’
                    _clear_right_border(mid_row_nxt, boundary_col - 1)

        # 3) 合计行：B79 总时间，C79 手作业时间，D79 自动时间，E79 步行时间
        if isinstance(totals, dict):
            total_sec = totals.get("total", 0.0)
            manual_sec = totals.get("manual", 0.0)
            auto_sec = totals.get("auto", 0.0)
            walk_sec = totals.get("walk", 0.0)
        else:
            total_sec = manual_sec = auto_sec = walk_sec = 0.0

        def _fmt_sec(v):
            """把秒数统一转成整数秒写入单元格"""
            try:
                return int(round(float(v)))
            except Exception:
                return v

        _set_value("B79", _fmt_sec(total_sec))   # 合计下面：总时间
        _set_value("C79", _fmt_sec(manual_sec))  # 手作业合计
        _set_value("D79", _fmt_sec(auto_sec))    # 自动合计
        _set_value("E79", _fmt_sec(walk_sec))    # 步行合计

        # 4) 在上方空白处写入工程信息
        _set_value("B2", project)
        _set_value("B3", part)
        _set_value("B4", worker)
        _set_value("E2", takt_sec)

        # 5) 保存
        wb.save(path)

    def export_single_placeholder(self):
        """
        单工程组合票导出流程：
        1. 读取 Tab2 中 A→B 区间作业数据并校验
        2. 选择保存路径
        3. 使用固定 Excel 模板导出标准作业组合票
        """
        try:
            project, part, worker, takt_sec, steps, totals = self._collect_single_inputs()
        except Exception as e:
            QMessageBox.warning(self, "输入有误", str(e))
            return

        default_name = f"{project}_单人组合票.xlsx" if project else "单人组合票.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出标准作业组合票",
            default_name,
            "Excel (*.xlsx)",
        )
        if not path:
            return

        try:
            self._export_single_to_excel(path, project, part, worker, takt_sec, steps, totals)
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
            return

        msg = (
            f"已导出标准作业组合票：\n{path}\n\n"
            f"工程名称：{project}\n"
            f"品番·品名：{part or '（未填写）'}\n"
            f"作业者：{worker or '（未填写）'}\n\n"
            f"节拍 TT：{takt_sec} 秒\n"
            f"总时间：{totals['total']:.1f} 秒\n"
            f"  其中 手作业：{totals['manual']:.1f} 秒\n"
            f"       自动：{totals['auto']:.1f} 秒\n"
            f"       步行：{totals['walk']:.1f} 秒\n\n"
            f"步骤数：{len(steps)} 步"
        )
        QMessageBox.information(self, "单工程组合票 - 导出完成", msg)

    # -------- 多车组合票：数据收集 & 导出 --------
    def fill_sample(self):
        """
        排程模型 v2 示例：设备数量 + 所属线别 + A/B/C 工时。
        工时 > 0 表示该车型经过该岗位；工时 = 0 表示该车型跳过该岗位。
        """
        self.tbl.setRowCount(0)
        sample_rows = [
            # 序号, 工程名称,   设备数量, 所属线别, 岗位设备,       A工时, B工时, C工时
            ("1",  "电集",       "2",    "双线",     "电集",         "100", "100", "100"),
            ("2",  "空悬+快充",  "1",    "1号线",    "空悬快充设备", "0",   "200", "0"),
            ("3",  "四轮定位",   "2",    "双线",     "四轮定位",     "110", "110", "110"),
        ]
        for row in sample_rows:
            self.add_row()
            r = self.tbl.rowCount() - 1

            self.tbl.setItem(r, 0, QTableWidgetItem(str(row[0])))
            self.tbl.setItem(r, 1, QTableWidgetItem(str(row[1])))

            device_count_widget = self.tbl.cellWidget(r, 2)
            if isinstance(device_count_widget, QComboBox):
                device_count_widget.setCurrentText(str(row[2]))

            line_scope_widget = self.tbl.cellWidget(r, 3)
            if isinstance(line_scope_widget, QComboBox):
                line_scope_widget.setCurrentText(str(row[3]))

            self.tbl.setItem(r, 4, QTableWidgetItem(str(row[4])))
            self.tbl.setItem(r, 5, QTableWidgetItem(str(row[5])))
            self.tbl.setItem(r, 6, QTableWidgetItem(str(row[6])))
            self.tbl.setItem(r, 7, QTableWidgetItem(str(row[7])))

        if not self.ed_project.text().strip():
            self.ed_project.setText("排程模型v2示例")
        self.spn_a_cars.setValue(4)
        self.spn_b_cars.setValue(2)
        self.spn_c_cars.setValue(0)
        self.spn_total_cars.setValue(60)
        self.cmb_grid.setCurrentText("1.0")
        self.cmb_wait.setCurrentText("开始前等待")
        self.cmb_launch_mode.setCurrentText("按数量投车")
        self.cmb_seq.setCurrentText("顺排(A→B→C)")
        self.spn_max_run.setValue(10)

    def _collect_multi_raw_inputs(self):
        """Collect multi-project UI values as plain data for later parser migration."""
        project = self.ed_project.text().strip() or "工程"
        cars_a = int(self.spn_a_cars.value())
        cars_b = int(self.spn_b_cars.value())
        cars_c = int(self.spn_c_cars.value())
        analysis_minutes = int(self.spn_total_cars.value()) if hasattr(self, "spn_total_cars") else 0
        target_takt = float(self.spn_target_takt.value()) if hasattr(self, "spn_target_takt") else 0.0
        is_ratio_mode = self.cmb_launch_mode.currentIndex() == 1
        sequence_mode_index = self.cmb_seq.currentIndex() if hasattr(self, "cmb_seq") else 0
        max_consecutive = int(self.spn_max_run.value()) if hasattr(self, "spn_max_run") else 10

        station_rows = []
        for r in range(self.tbl.rowCount()):
            seq = (self.tbl.item(r, 0).text().strip() if self.tbl.item(r, 0) else "")
            name = (self.tbl.item(r, 1).text().strip() if self.tbl.item(r, 1) else "")

            device_count_widget = self.tbl.cellWidget(r, 2)
            device_count_text = device_count_widget.currentText().strip() if isinstance(device_count_widget, QComboBox) else "2"

            line_scope_widget = self.tbl.cellWidget(r, 3)
            line_scope = line_scope_widget.currentText().strip() if isinstance(line_scope_widget, QComboBox) else "双线"

            grp = (self.tbl.item(r, 4).text().strip() if self.tbl.item(r, 4) else "")
            dur_a = (self.tbl.item(r, 5).text().strip() if self.tbl.item(r, 5) else "")
            dur_b = (self.tbl.item(r, 6).text().strip() if self.tbl.item(r, 6) else "")
            dur_c = (self.tbl.item(r, 7).text().strip() if self.tbl.item(r, 7) else "")

            station_rows.append({
                "seq": seq,
                "display": name,
                "device_count": device_count_text,
                "line_scope": line_scope,
                "group": grp,
                "duration_a": dur_a,
                "duration_b": dur_b,
                "duration_c": dur_c,
                "color": "",
            })

        return {
            "project": project,
            "cars_a": cars_a,
            "cars_b": cars_b,
            "cars_c": cars_c,
            "analysis_minutes": analysis_minutes,
            "target_takt": target_takt,
            "is_ratio_mode": is_ratio_mode,
            "sequence_mode_index": sequence_mode_index,
            "max_consecutive": max_consecutive,
            "station_rows": station_rows,
        }

    def _parse_multi_inputs_from_raw(self):
        """
        通过 core.input_parser 解析多工程原始输入。
        当前阶段仅作为后续替换 _collect_inputs() 的桥接函数。
        暂不接入现有分析/导出流程。
        """
        raw_inputs = self._collect_multi_raw_inputs()
        return core_parse_multi_project_inputs(raw_inputs)

    def _collect_inputs_from_parser_tuple(self):
        """
        使用 core.input_parser 解析多工程输入，并转换为 _collect_inputs() 兼容的旧 tuple。
        当前阶段仅作为后续替换 _collect_inputs() 的桥接函数。
        暂不接入现有分析/导出流程。
        """
        parsed = self._parse_multi_inputs_from_raw()

        self.current_analysis_time_seconds = parsed.get("analysis_time_seconds")
        self.current_theoretical_launch_count = parsed.get("theoretical_launch_count")

        return (
            parsed["project"],
            parsed["cars"],
            parsed["grid_step"],
            parsed["wait_policy"],
            parsed["defs"],
            parsed["vehicle_counts"],
            parsed["sequence_mode"],
            parsed["max_consecutive"],
            parsed["ratio_pattern"],
            parsed["target_takt"],
        )

    def _collect_inputs(self):
        return self._collect_inputs_from_parser_tuple()

    def do_export(self):
        try:
            project, cars, grid_step, wait_policy, defs, vehicle_counts, sequence_mode, max_consecutive, ratio_pattern, target_takt = self._collect_inputs()
        except Exception as e:
            QMessageBox.warning(self, "输入有误", str(e))
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "导出位置",
            f"{project}_组合票.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        self.dst_path = path

        worker = Worker(
            tickets.schedule_and_export,
            defs, cars, grid_step, wait_policy, project, self.dst_path,
            vehicle_counts, sequence_mode, max_consecutive, ratio_pattern, target_takt,
        )
        worker.signals.error.connect(self._on_error)
        worker.signals.finished.connect(self._on_export_finished)
        self.thread_pool.start(worker)
        self.status.showMessage("正在生成组合票...", 5000)

    def _on_export_finished(self, *args):
        self.status.showMessage("导出完成", 6000)
        QMessageBox.information(self, "完成", f"已导出：\n{self.dst_path}")

    def _fmt_analysis_num(self, value):
        try:
            v = float(value)
        except Exception:
            return str(value)
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:.1f}"
    
    def _apply_time_window_analysis(self, analysis, rows, target_takt):
        """v2-6D：按比例投车的时间窗口产能分析薄封装。"""
        return core_apply_time_window_analysis(
            analysis=analysis,
            rows=rows,
            target_takt=target_takt,
            analysis_time_seconds=getattr(self, "current_analysis_time_seconds", None),
            theoretical_launch_count=getattr(self, "current_theoretical_launch_count", None),
        )

    def _show_analysis_result(self, analysis):
        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}

        total_cars = summary.get("total_cars", 0)
        max_finish = self._fmt_analysis_num(summary.get("max_finish", 0.0))
        total_wait = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        avg_wait = self._fmt_analysis_num(summary.get("avg_wait", 0.0))
        blocking_result = summary.get("blocking_result", "无阻塞") or "无阻塞"
        blocking_time = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        overflow_vehicle_count = self._fmt_analysis_num(summary.get("overflow_vehicle_count", 0.0))
        blocking_station_text = summary.get("blocking_station_text", "无") or "无"
        batch_overrun_time = self._fmt_analysis_num(summary.get("batch_overrun_time", 0.0))
        batch_overrun_cars = self._fmt_analysis_num(summary.get("batch_overrun_cars", 0.0))
        process_root_text = summary.get("process_over_takt_root_text", "无") or "无"
        takt_result = summary.get("takt_result", "未设定") or "未设定"
        over_count = summary.get("over_takt_station_count", 0)
        process_root_text = summary.get("process_over_takt_root_text", "无") or "无"
        batch_overrun_raw = float(summary.get("batch_overrun_time", 0.0) or 0.0)
        final_result = "NG" if batch_overrun_raw > 0 else "OK"

        text = (
            "结果分析："
            f"模型最终判定 {final_result} ｜ "
            f"累计阻塞 {blocking_time} 秒 ｜ "
            f"溢出工时 {batch_overrun_time} 秒 / {batch_overrun_cars} 台 ｜ "
            f"超节拍工程 {process_root_text}"
        )


        if hasattr(self, "lbl_analysis"):
            self.lbl_analysis.setText(text)
        else:
            QMessageBox.information(self, "排程分析完成", text)

        if hasattr(self, "_show_station_analysis"):
            self._show_station_analysis(analysis)

        if hasattr(self, "_show_vehicle_summary"):
            self._show_vehicle_summary(analysis)


    def _on_error(self, err_msg):
        self.status.showMessage("导出失败", 6000)
        QMessageBox.critical(self, "导出失败", str(err_msg))


    # ---------- 帮助弹窗 ----------
    def show_help(self):
        msg = (
            "<h3>组合票操作指南</h3>"
            "<ol>"
            "<li>多工程组合票按『设备数量 + 所属线别 + 岗位设备 + A/B/C 工时』录入</li>"
            "<li>设备数量可选：1 / 2；设备数量为 2 时，所属线别固定为『双线』</li>"
            "<li>设备数量为 1 时，所属线别可选：1号线 / 2号线 / 双线共用</li>"
            "<li>A/B/C 工时大于 0 表示该车型经过该岗位；工时为 0 表示该车型跳过该岗位</li>"
            "<li>参与投车的车型，工时不能为空；未参与投车的车型，工时可以为空</li>"
            "<li>投车模式支持：按数量投车 / 按比例投车；按数量投车下可选择顺排(A→B→C)或交替混流</li>"
            "<li>顺排/交替混流模式下，A/B/C 填数量；按比例投车模式下，A/B/C 填比例，并用分析时间与目标节拍计算理论投车台数</li>"
            "<li>最大连续台数默认 10；填 1 表示尽量强制交替</li>"
            "<li>填写完点击『分析当前排程』可查看结果；点击『生成并导出组合票』即可生成 Excel</li>"
            "</ol>"
        )
        QMessageBox.information(self, "帮助", msg)


    def _show_station_analysis(self, analysis):
        if not hasattr(self, "tbl_station_analysis"):
            return

        station_stats = analysis.get("station_stats", []) if isinstance(analysis, dict) else []
        self.tbl_station_analysis.setRowCount(0)

        for item in station_stats:
            r = self.tbl_station_analysis.rowCount()
            self.tbl_station_analysis.insertRow(r)

            values = [
                str(item.get("station", "")),
                str(item.get("count", 0)),
                self._fmt_analysis_num(item.get("total_process", 0.0)),
                self._fmt_analysis_num(item.get("blocking_time", item.get("overflow_wait_time", item.get("total_block_wait", 0.0)))),
                self._fmt_analysis_num(item.get("avg_process", 0.0)),
                self._fmt_analysis_num(item.get("avg_overflow_wait", item.get("avg_block_wait", 0.0))),
                str(item.get("takt_result", "未设定")),
                str(item.get("over_takt_types", "—")),
            ]

            for c, value in enumerate(values):
                table_item = QTableWidgetItem(value)
                if c > 0:
                    table_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    table_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                self.tbl_station_analysis.setItem(r, c, table_item)

        # 固定列宽，不再自动根据内容调整，避免每次分析后表格宽度跳动
        # self.tbl_station_analysis.resizeColumnsToContents()
    def _show_vehicle_summary(self, analysis):
        """在车型数据摘要区展示当前排程的车型构成与关键结果。"""
        if not hasattr(self, "lbl_vehicle_summary"):
            return

        summary = analysis.get("summary", {}) if isinstance(analysis, dict) else {}
        car_stats = analysis.get("car_stats", []) if isinstance(analysis, dict) else []
        car_type_summary = analysis.get("car_type_summary", analysis.get("type_stats", [])) if isinstance(analysis, dict) else []

        type_counts = {"A": 0, "B": 0, "C": 0}
        if car_stats:
            for item in car_stats:
                vt = str(item.get("car_type", "") or "").strip().upper()
                if vt in type_counts:
                    type_counts[vt] += 1
        else:
            for item in car_type_summary:
                vt = str(item.get("car_type", "") or "").strip().upper()
                if vt in type_counts:
                    type_counts[vt] += int(item.get("count", 0) or 0)

        total_cars = int(summary.get("total_cars", 0) or 0)
        if total_cars <= 0:
            total_cars = sum(type_counts.values())

        # v2-5 blocking analysis fields
        target_takt = self._fmt_analysis_num(summary.get("target_takt", 0.0))
        max_finish = self._fmt_analysis_num(summary.get("max_finish", 0.0))
        total_wait = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        avg_wait = self._fmt_analysis_num(summary.get("avg_wait", 0.0))
        takt_result = summary.get("takt_result", "未设定") or "未设定"
        over_count = summary.get("over_takt_station_count", 0)
        blocking_result = summary.get("blocking_result", "无阻塞") or "无阻塞"
        blocking_station_text = summary.get("blocking_station_text", "无") or "无"
        blocking_station_count = summary.get("blocking_station_count", 0)
        blocking_time = self._fmt_analysis_num(summary.get("total_wait", 0.0))
        overflow_vehicle_count = self._fmt_analysis_num(summary.get("overflow_vehicle_count", 0.0))
        batch_overrun_time = self._fmt_analysis_num(summary.get("batch_overrun_time", 0.0))
        batch_overrun_cars = self._fmt_analysis_num(summary.get("batch_overrun_cars", 0.0))
        process_root_text = summary.get("process_over_takt_root_text", "无") or "无"

        batch_overrun_raw = float(summary.get("batch_overrun_time", 0.0) or 0.0)

        time_window_result = summary.get("time_window_result", "") or ""
        final_result = time_window_result or "OK"

        actual_output_count_raw = int(summary.get("actual_output_count_in_window", 0) or 0)
        display_actual_output_count = self._fmt_analysis_num(summary.get(
            "display_actual_output_count_in_window",
            actual_output_count_raw,
        ))
        planned_output_count = self._fmt_analysis_num(summary.get("planned_output_count_in_window", 0.0))
        theoretical_launch_count = int(summary.get("theoretical_launch_count", total_cars) or total_cars)
        achievement_rate = float(summary.get("achievement_rate", 0.0) or 0.0) * 100
        target_takt_display = self._fmt_analysis_num(summary.get("target_takt", self.spn_target_takt.value() if hasattr(self, "spn_target_takt") else 0))

        def _fmt_optional_seconds(value):
            if value is None:
                return "—"
            return self._fmt_analysis_num(value)

        planned_n_finish_time = _fmt_optional_seconds(summary.get("planned_n_finish_time"))
        actual_n_finish_time = _fmt_optional_seconds(summary.get("actual_n_finish_time"))
        finish_delta_raw = summary.get("finish_delta")
        finish_delta_num = None
        if finish_delta_raw is None:
            finish_delta = "—"
        else:
            try:
                finish_delta_num = float(finish_delta_raw)
                finish_delta = f"{finish_delta_num:+.1f}" if abs(finish_delta_num - round(finish_delta_num)) >= 1e-9 else f"{int(round(finish_delta_num)):+d}"
            except Exception:
                finish_delta = str(finish_delta_raw)
        actual_line_takt_in_window = _fmt_optional_seconds(summary.get("actual_line_takt_in_window"))

        def _metric_card(title, value, value_color="#0f172a"):
            return (
                "<td width='20%' style='"
                "width:20%;"
                "border:1px solid #dbe3ef;"
                "background:#ffffff;"
                "padding:8px 9px;"
                "vertical-align:top;"
                "'>"
                f"<div style='font-size:10px;color:#64748b;'>{title}</div>"
                f"<div style='font-size:14px;font-weight:700;color:{value_color};margin-top:3px;'>{value}</div>"
                "</td>"
            )

        is_ratio = hasattr(self, "cmb_launch_mode") and self.cmb_launch_mode.currentIndex() == 1
        if is_ratio:
            ratio_a = int(self.spn_a_cars.value()) if hasattr(self, "spn_a_cars") else 0
            ratio_b = int(self.spn_b_cars.value()) if hasattr(self, "spn_b_cars") else 0
            ratio_c = int(self.spn_c_cars.value()) if hasattr(self, "spn_c_cars") else 0
            analysis_minutes = int(self.spn_total_cars.value()) if hasattr(self, "spn_total_cars") else 0
            summary_lines = [
                (
                    ("模式", "按比例投车"),
                    ("比例", f"{ratio_a}:{ratio_b}:{ratio_c}"),
                ),
                (
                    ("时间", f"{analysis_minutes}分钟"),
                    ("目标节拍", f"{target_takt_display}s"),
                ),
                (
                    ("投车台数", f"{total_cars}台"),
                ),
            ]
        else:
            seq_text = self.cmb_seq.currentText() if hasattr(self, "cmb_seq") else "—"
            summary_lines = [
                (
                    ("模式", "按数量投车"),
                    ("A/B/C", f"A{type_counts['A']}/B{type_counts['B']}/C{type_counts['C']}"),
                ),
                (
                    ("排列", seq_text),
                    ("目标节拍", f"{target_takt_display}s"),
                ),
                (
                    ("投车台数", f"{total_cars}台"),
                ),
            ]

        def _summary_line(items):
            return "｜".join(
                f"<span style='color:#64748b;'>{label}：</span><b style='color:#0f172a;'>{value}</b>"
                for label, value in items
            )

        left_html = (
            "<div style='font-weight:700;color:#334155;margin-bottom:6px;'>基础排程摘要</div>"
            "<div style='font-size:13px;color:#1e293b;line-height:1.75;'>"
            + "<br>".join(_summary_line(line) for line in summary_lines)
            + "</div>"
        )
        if is_ratio and time_window_result:
            if str(final_result).upper() == "OK":
                final_color = "#16a34a"
            elif str(final_result).upper() == "NG":
                final_color = "#dc2626"
            else:
                final_color = "#0f172a"

            if finish_delta_num is None:
                finish_delta_color = "#0f172a"
                remark_delta = "差异—"
            elif finish_delta_num > 0:
                finish_delta_color = "#dc2626"
                remark_delta = f"超时{self._fmt_analysis_num(abs(finish_delta_num))}s"
            elif finish_delta_num < 0:
                finish_delta_color = "#16a34a"
                remark_delta = f"富余{self._fmt_analysis_num(abs(finish_delta_num))}s"
            else:
                finish_delta_color = "#0f172a"
                remark_delta = "准时0s"

            remark_color = finish_delta_color
            model_cards = [
                _metric_card("最终判定", final_result, final_color),
                _metric_card("下线台数", f"{display_actual_output_count}/{planned_output_count}"),
                _metric_card("达成率", f"{achievement_rate:.1f}%"),
                _metric_card("实际节拍", f"{actual_line_takt_in_window}/{target_takt_display}"),
                _metric_card("累计阻塞", f"{blocking_time}s"),
            ]
            right_html = (
                "<div style='font-size:13px;font-weight:700;color:#334155;margin-bottom:4px;'>模型结果</div>"
                "<div style='margin-bottom:2px;'>"
                "<table width='100%' cellspacing='2' cellpadding='0' style='width:100%;'>"
                "<tr>"
                + "".join(model_cards)
                + "</tr></table>"
                "</div>"
                f"<div style='font-size:11px;color:#334155;line-height:1.3;margin-top:2px;margin-bottom:0;'>超节拍工程：{process_root_text}｜备注：计划{planned_n_finish_time}s完成｜实际{actual_n_finish_time}s完成｜"
                f"<span style='color:{remark_color};font-weight:700;'>{remark_delta}</span></div>"
            )
        else:
            try:
                target_takt_raw = float(summary.get(
                    "target_takt",
                    self.spn_target_takt.value() if hasattr(self, "spn_target_takt") else 0.0
                ) or 0.0)
            except Exception:
                target_takt_raw = 0.0

            try:
                max_finish_raw = float(summary.get("max_finish", 0.0) or 0.0)
            except Exception:
                max_finish_raw = 0.0

            try:
                total_wait_raw = float(summary.get("total_wait", 0.0) or 0.0)
            except Exception:
                total_wait_raw = 0.0

            station_count = 0
            try:
                station_count = len(self._sim_station_names())
            except Exception:
                station_count = 0

            if station_count <= 0:
                station_stats = analysis.get("station_stats", []) if isinstance(analysis, dict) else []
                station_count = len(station_stats)

            station_count = max(1, station_count)

            if target_takt_raw > 0 and total_cars > 0:
                planned_finish_raw = (total_cars + station_count - 1) * target_takt_raw
            else:
                planned_finish_raw = 0.0

            planned_finish = self._fmt_analysis_num(planned_finish_raw)

            has_over_takt_root = process_root_text not in ("无", "", "—")
            has_finish_gap = planned_finish_raw > 0 and max_finish_raw > planned_finish_raw + 1e-6
            has_blocking = total_wait_raw > 0

            quantity_final_result = "NG" if (has_finish_gap or has_blocking or has_over_takt_root) else "OK"
            quantity_final_color = "#dc2626" if quantity_final_result == "NG" else "#16a34a"

            model_cards = [
                _metric_card("最终判定", quantity_final_result, quantity_final_color),
                _metric_card("生成台数", f"{total_cars}台"),
                _metric_card("完成时间", f"{max_finish}s"),
                _metric_card("计划时间", f"{planned_finish}s"),
                _metric_card("累计阻塞", f"{blocking_time}s"),
            ]
            right_html = (
                "<div style='font-size:13px;font-weight:700;color:#334155;margin-bottom:4px;'>模型结果</div>"
                "<div style='margin-bottom:2px;'>"
                "<table width='100%' cellspacing='2' cellpadding='0' style='width:100%;'>"
                "<tr>"
                + "".join(model_cards)
                + "</tr></table>"
                "</div>"
                f"<div style='font-size:11px;color:#334155;line-height:1.3;margin-top:2px;margin-bottom:0;'>超节拍工程：{process_root_text}</div>"
            )
        
        html = (
            "<table width='100%' cellspacing='0' cellpadding='0'>"
            "<tr>"
            f"<td width='30%' valign='top' style='width:30%;padding-right:12px;border-right:1px solid #dbe3ef;'>{left_html}</td>"
            f"<td width='70%' valign='top' style='width:70%;padding-left:12px;'>{right_html}</td>"
            "</tr>"
            "</table>"
        )
        self.lbl_vehicle_summary.setText(html)
    def _active_sim_rows(self):
        """返回当前仿真时间正在加工的排程段。"""
        rows = getattr(self, "last_schedule_rows", []) or []
        current = float(getattr(self, "sim_time", 0.0) or 0.0)
        active = []
        for row in rows:
            start = self._sim_row_start(row)
            end = self._sim_row_end(row)
            if start <= current < end:
                active.append((start, end, row))
        active.sort(key=lambda x: (self._sim_row_station(x[2]), x[0]))
        return active

    def _draw_sim_scene(self):
        """绘制横向双轨坐标系的生产线仿真底图。"""
        if not hasattr(self, "sim_scene"):
            return

        self.sim_scene.clear()
        rows = getattr(self, "last_schedule_rows", []) or []
        if not rows:
            self.sim_scene.addText("请先点击『分析当前排程』生成仿真数据。")
            return

        station_defs = self._sim_station_defs()
        station_names = [item["name"] for item in station_defs]
        if not station_names:
            self.sim_scene.addText("暂无岗位数据。")
            return

        station_count = len(station_names)
        view_w = 0
        if hasattr(self, "sim_graphics_view") and self.sim_graphics_view is not None:
            try:
                view_w = int(self.sim_graphics_view.viewport().width() or 0)
            except Exception:
                view_w = 0
        view_w = max(920, view_w or 0)

        margin_x = 20
        line_label_w = 64
        track_lead_w = 14
        zone_top = 24
        line1_y = 88
        line2_y = 148
        lane_gap = line2_y - line1_y
        title_band_h = 34

        def _get_station_metrics(count, available_width):
            if count <= 4:
                pref_station_w, pref_gap = 208, 24
            elif count == 5:
                pref_station_w, pref_gap = 176, 14
            elif count == 6:
                pref_station_w, pref_gap = 154, 12
            elif count == 7:
                pref_station_w, pref_gap = 134, 10
            else:
                pref_station_w, pref_gap = 118, 8

            if count <= 1:
                return pref_station_w, pref_gap

            pref_total = count * pref_station_w + (count - 1) * pref_gap
            if pref_total <= available_width:
                return pref_station_w, pref_gap

            min_gap = 6
            min_station_w = 96 if count >= 8 else 108
            fit_station_w = int((available_width - (count - 1) * min_gap) / max(1, count))
            fit_station_w = max(min_station_w, fit_station_w)
            return fit_station_w, min_gap

        usable_track_w = max(620, view_w - margin_x * 2 - line_label_w - track_lead_w - 28)
        station_w, station_gap = _get_station_metrics(station_count, usable_track_w)
        compact_mode = station_w <= 134
        slot_h = 38 if compact_mode else 40
        slot_w = max(82, station_w - 24)
        zone_h = int((line2_y + slot_h / 2 + 18) - zone_top)
        track_start_x = margin_x + line_label_w
        first_station_x = track_start_x + track_lead_w
        last_station_x = first_station_x + (station_count - 1) * (station_w + station_gap)
        track_end_x = last_station_x + station_w + 18
        scene_w = max(view_w - 6, track_end_x + margin_x)
        scene_h = zone_top + zone_h + 14

        self.sim_scene.addRect(
            0,
            0,
            scene_w,
            scene_h,
            QPen(QColor("#243244")),
            QBrush(QColor("#243244")),
        )

        def _station_device_mode(station_info):
            scope = str(station_info.get("line_scope") or "").strip()
            try:
                device_count = int(float(station_info.get("device_count") or 0))
            except Exception:
                device_count = 0

            if scope in ("1号线", "2号线"):
                return "single"
            if scope == "双线共用":
                return "shared"
            if scope == "双线" and device_count == 1:
                return "shared"
            return "parallel"

        def _line_slots(line_scope):
            scope = str(line_scope or "").strip()
            if scope == "1号线":
                return [("1号线", True), ("2号线", False)]
            if scope == "2号线":
                return [("1号线", False), ("2号线", True)]
            return [("1号线", True), ("2号线", True)]

        def _line_key(line_no):
            text = str(line_no or "").strip()
            if "2" in text:
                return "2号线"
            return "1号线"

        def _build_sim_track_layout():
            layout = {}
            for idx, station_info in enumerate(station_defs):
                x = first_station_x + idx * (station_w + station_gap)
                slot_x = x + max(0, (station_w - slot_w) / 2)
                layout[station_info["name"]] = {
                    "x": x,
                    "title_x": x + 12,
                    "slot_x": slot_x,
                    "1号线": {
                        "x": slot_x,
                        "y": line1_y - slot_h / 2,
                        "w": slot_w,
                        "h": slot_h,
                    },
                    "2号线": {
                        "x": slot_x,
                        "y": line2_y - slot_h / 2,
                        "w": slot_w,
                        "h": slot_h,
                    },
                }
            return layout

        def _draw_track_headers():
            for label, y in (("1号线", line1_y), ("2号线", line2_y)):
                head_item = self.sim_scene.addText(label)
                head_item.setDefaultTextColor(QColor("#cbd5e1"))
                font = QFont()
                font.setPointSize(11)
                font.setBold(True)
                head_item.setFont(font)
                head_item.setPos(margin_x, y - 15)
                self.sim_scene.addLine(
                    track_start_x - 12,
                    y,
                    track_start_x - 1,
                    y,
                    QPen(QColor("#64748b"), 1),
                )

        def _draw_lane_lines():
            channel_h = 22
            channel_fill = QBrush(QColor("#334155"))
            channel_border = QPen(QColor(100, 116, 139, 150), 1)
            texture_pen = QPen(QColor(71, 85, 105, 115), 1)
            for idx in range(station_count - 1):
                left_x = first_station_x + idx * (station_w + station_gap) + station_w - 1
                right_x = first_station_x + (idx + 1) * (station_w + station_gap) + 1
                if right_x <= left_x:
                    continue
                for y in (line1_y, line2_y):
                    self.sim_scene.addRect(
                        left_x,
                        y - channel_h / 2,
                        right_x - left_x,
                        channel_h,
                        QPen(Qt.NoPen),
                        channel_fill,
                    )
                    self.sim_scene.addLine(
                        left_x,
                        y - channel_h / 2,
                        right_x,
                        y - channel_h / 2,
                        channel_border,
                    )
                    self.sim_scene.addLine(
                        left_x,
                        y + channel_h / 2,
                        right_x,
                        y + channel_h / 2,
                        channel_border,
                    )
                    texture_x = left_x + 8
                    while texture_x < right_x - 4:
                        self.sim_scene.addLine(
                            texture_x,
                            y - channel_h / 2 + 3,
                            texture_x,
                            y + channel_h / 2 - 3,
                            texture_pen,
                        )
                        texture_x += 10

        def _draw_station_zone(station_info, layout_info):
            x = layout_info["x"]
            self.sim_scene.addRect(
                x,
                zone_top,
                station_w,
                zone_h,
                QPen(QColor(148, 163, 184, 105), 1),
                QBrush(QColor(203, 213, 225, 90)),
            )

            title_item = self.sim_scene.addText(f"ST-{station_info['seq']} {station_info['name']}")
            title_item.setDefaultTextColor(QColor("#e5e7eb"))
            title_font = QFont()
            title_font.setPointSize(9 if station_count >= 8 else (10 if compact_mode else 11))
            title_font.setBold(True)
            title_item.setFont(title_font)
            title_item.document().setDocumentMargin(0)
            title_item.setTextWidth(max(74, station_w - 20))
            title_item.setPos(layout_info["title_x"] - 2, zone_top + 11)

            if _station_device_mode(station_info) == "shared":
                badge_w = 56
                badge_h = 18
                badge_x = x + station_w - badge_w - 12
                badge_y = zone_top + 12
                self.sim_scene.addRect(
                    badge_x,
                    badge_y,
                    badge_w,
                    badge_h,
                    QPen(QColor("#f59e0b"), 1),
                    QBrush(QColor("#fffbeb")),
                )
                badge_item = self.sim_scene.addText("共用设备")
                badge_item.setDefaultTextColor(QColor("#b45309"))
                badge_item.setPos(badge_x + 6, badge_y + 1)

        def _draw_station_track_slot(slot, enabled):
            if enabled:
                pen = QPen(QColor(148, 163, 184, 95), 1)
                brush = QBrush(QColor(248, 250, 252, 175))
            else:
                pen = QPen(QColor(156, 163, 175, 75), 1, Qt.DashLine)
                brush = QBrush(QColor(203, 213, 225, 70))
            self.sim_scene.addRect(
                slot["x"],
                slot["y"],
                slot["w"],
                slot["h"],
                pen,
                brush,
            )
            status_text = "空闲" if enabled else "不适用"
            status_item = self.sim_scene.addText(status_text)
            slot_font = QFont()
            slot_font.setPointSize(9 if compact_mode else 10)
            status_item.setFont(slot_font)
            status_item.setDefaultTextColor(QColor("#94a3b8" if enabled else "#9ca3af"))
            status_item.setPos(slot["x"] + 10, slot["y"] + max(8, (slot["h"] - 18) / 2))

        def _draw_slot_waiting_outline(slot):
            self.sim_scene.addRect(
                slot["x"],
                slot["y"],
                slot["w"],
                slot["h"],
                QPen(QColor("#f97316"), 2),
            )

        def _draw_overflow_badge(x, y, count):
            badge_w = 52
            badge_h = 18
            self.sim_scene.addRect(
                x,
                y,
                badge_w,
                badge_h,
                QPen(QColor("#64748b"), 1),
                QBrush(QColor("#f8fafc")),
            )
            badge_item = self.sim_scene.addText(f"还有{count}台")
            badge_item.setDefaultTextColor(QColor("#334155"))
            badge_item.setPos(x + 4, y + 1)

        def _vehicle_block_label(row):
            car_type = str(
                self._sim_row_value(row, "car_type", "type", "vehicle_type", default="")
                or ""
            ).upper()
            car_no = str(self._sim_row_value(row, "car", "car_no", "car_id", default="") or "")
            return f"{car_type or '车'}-{car_no or '?'}"

        def _draw_vehicle_capsule(x, y, w, h, row, status, seconds_text, over_takt=False):
            car_type = str(
                self._sim_row_value(row, "car_type", "type", "vehicle_type", default="")
                or ""
            ).upper()
            if over_takt:
                border = QColor("#dc2626")
                fill = QColor("#fee2e2")
                status_color = QColor("#dc2626")
            elif status == "等待":
                border = QColor("#f97316")
                fill = QColor("#ffedd5")
                status_color = QColor("#c2410c")
            else:
                border = QColor("#16a34a")
                fill = QColor("#dcfce7")
                status_color = QColor("#15803d")

            if car_type == "A":
                type_bar = QColor("#2563eb")
            elif car_type == "B":
                type_bar = QColor("#ea580c")
            elif car_type == "C":
                type_bar = QColor("#16a34a")
            else:
                type_bar = QColor("#64748b")

            self.sim_scene.addRect(
                x,
                y,
                w,
                h,
                QPen(border, 2),
                QBrush(fill),
            )
            self.sim_scene.addRect(
                x,
                y,
                5,
                h,
                QPen(type_bar, 1),
                QBrush(type_bar),
            )
            label_item = self.sim_scene.addText(f"{_vehicle_block_label(row)} {status}")
            label_font = QFont()
            label_font.setPointSize(9 if compact_mode else 10)
            label_font.setBold(True)
            label_item.setFont(label_font)
            label_item.setDefaultTextColor(status_color)
            label_item.document().setDocumentMargin(0)
            label_item.setTextWidth(max(58, w - 16))
            label_item.setPos(x + 10, y + 4)
            detail_item = self.sim_scene.addText(seconds_text)
            detail_font = QFont()
            detail_font.setPointSize(8 if compact_mode else 9)
            detail_item.setFont(detail_font)
            detail_item.setDefaultTextColor(QColor("#334155"))
            detail_item.document().setDocumentMargin(0)
            detail_item.setTextWidth(max(58, w - 16))
            detail_item.setPos(x + 10, y + (20 if compact_mode else 21))

        station_layout = _build_sim_track_layout()
        _draw_track_headers()
        _draw_lane_lines()

        slot_rects = {}
        station_modes = {}
        for station_info in station_defs:
            name = station_info["name"]
            line_scope = station_info.get("line_scope")
            station_modes[name] = _station_device_mode(station_info)
            layout_info = station_layout[name]
            _draw_station_zone(station_info, layout_info)
            slots = _line_slots(line_scope)
            slot_rects[name] = {}
            for line_label, enabled in slots:
                slot = dict(layout_info[line_label])
                slot["enabled"] = enabled
                _draw_station_track_slot(slot, enabled)
                slot_rects[name][line_label] = slot

        current = float(getattr(self, "sim_time", 0.0) or 0.0)
        target_takt = float(self.spn_target_takt.value()) if hasattr(self, "spn_target_takt") else 0.0
        vehicle_blocks = []

        for row in rows:
            start = self._sim_row_start(row)
            svc_finish = self._sim_row_service_finish(row)
            if start <= current < svc_finish:
                elapsed = max(0.0, current - start)
                remain = max(0.0, svc_finish - current)
                vehicle_blocks.append({
                    "station": self._sim_row_station(row),
                    "line": _line_key(self._sim_row_line_no(row)),
                    "row": row,
                    "status": "超时" if target_takt > 0 and elapsed > target_takt else "加工",
                    "seconds": f"超时{elapsed - target_takt:.1f}s" if target_takt > 0 and elapsed > target_takt else f"剩余{remain:.1f}s",
                    "over_takt": target_takt > 0 and elapsed > target_takt,
                    "order_time": start,
                })

        car_rows = self._sim_car_rows()
        for _, car_segments in car_rows.items():
            if not car_segments:
                continue
            for idx, seg in enumerate(car_segments):
                next_seg = car_segments[idx + 1] if idx + 1 < len(car_segments) else None
                wait_start = self._sim_row_service_finish(seg)
                wait_end = self._sim_row_wait_end(seg, next_seg)
                if wait_start <= current < wait_end:
                    vehicle_blocks.append({
                        "station": self._sim_row_station(seg),
                        "line": _line_key(self._sim_row_line_no(seg)),
                        "row": seg,
                        "status": "等待",
                        "seconds": self._sim_wait_label(seg, next_seg, current, wait_end, set()),
                        "over_takt": False,
                        "order_time": wait_start,
                    })
                    break

        shared_processing = {}
        for item in sorted(vehicle_blocks, key=lambda value: value["order_time"]):
            if station_modes.get(item["station"]) != "shared":
                continue
            if item["status"] == "等待":
                continue
            station_name = item["station"]
            if station_name not in shared_processing:
                shared_processing[station_name] = item
                continue
            item["status"] = "等待"
            item["seconds"] = "等待"
            item["over_takt"] = False

        def _vehicle_sort_key(item):
            station_index = station_names.index(item["station"]) if item["station"] in station_names else 999
            status_rank = 1 if item["status"] == "等待" else 0
            return station_index, item["line"], status_rank, item["order_time"]

        vehicle_blocks.sort(key=_vehicle_sort_key)
        slot_groups = {}
        for item in vehicle_blocks:
            station_slots = slot_rects.get(item["station"])
            if not station_slots:
                continue
            slot_line = item["line"]
            slot = station_slots.get(item["line"])
            if not slot or not slot.get("enabled"):
                fallback = next(
                    (
                        (line_label, candidate)
                        for line_label, candidate in station_slots.items()
                        if candidate.get("enabled")
                    ),
                    None,
                )
                if fallback:
                    slot_line, slot = fallback
            if not slot:
                continue
            slot_key = (item["station"], slot_line)
            slot_groups.setdefault(slot_key, {"slot": slot, "items": []})["items"].append(item)

        for slot_group in slot_groups.values():
            slot = slot_group["slot"]
            items = slot_group["items"]
            if not items:
                continue
            main_item = items[0]
            overflow_count = max(0, len(items) - 1)
            block_w = slot["w"] - 8
            if overflow_count:
                block_w = max(82, block_w - 58)
            _draw_vehicle_capsule(
                slot["x"] + 4,
                slot["y"] + 3,
                block_w,
                slot["h"] - 6,
                main_item["row"],
                main_item["status"],
                main_item["seconds"],
                main_item.get("over_takt", False),
            )
            if overflow_count:
                _draw_overflow_badge(
                    slot["x"] + slot["w"] - 56,
                    slot["y"] + max(4, (slot["h"] - 18) / 2),
                    overflow_count,
                )

        self.sim_scene.setSceneRect(0, 0, scene_w, scene_h)
